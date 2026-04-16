#!/usr/bin/env python3
"""
link-auditor.py — WordPress Site Link Auditor

Connects to Bluehost server2 via SSH, lists all WordPress sites,
then for a chosen site:
  1. Table 1: All posts/pages with their live URLs
  2. Table 2: Per-post non-affiliate link audit (vs local MD source files)
     - Checks links are relative paths (not absolute)
     - Cross-verifies link targets match the source MD file
  3. Anomaly report: anything suspicious beyond link issues
"""

import subprocess
import sys
import re
import os
from pathlib import Path
from html.parser import HTMLParser
from urllib.parse import urlparse

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─── SSH CONFIG (from server2-exploration.md) ────────────────────────────────
SSH_HOST = "50.6.155.174"
SSH_USER = "dpskbcmy"
SSH_KEY  = os.path.expanduser("~/.ssh/id_rsa_bluehost2")
WP_ROOT  = f"/home1/{SSH_USER}/public_html"

# ─── Affiliate link domains to skip ──────────────────────────────────────────
AFFILIATE_DOMAINS = {"getyourguide.com", "tiqets.com", "viator.com", "gyg.eu"}

# ─── Local input folder base ─────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).resolve().parent.parent          # auto-create-site/
INPUT_BASE   = SCRIPT_DIR / "input"

# ─── Site → local folder (populated at runtime from CLI args) ────────────────
SITE_TO_INPUT: dict[str, str] = {}   # filled in main() from argv


# ─── SSH helper ──────────────────────────────────────────────────────────────

def ssh(cmd: str) -> str:
    result = subprocess.run(
        ["ssh", "-i", SSH_KEY, "-o", "StrictHostKeyChecking=no",
         "-o", "ConnectTimeout=15", f"{SSH_USER}@{SSH_HOST}", cmd],
        capture_output=True, text=True
    )
    if result.returncode != 0 and result.stderr.strip():
        # non-fatal — wp-cli sometimes writes warnings to stderr
        pass
    return result.stdout.strip()


# ─── HTML link extractor ──────────────────────────────────────────────────────

class LinkExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.links = []  # list of (href, anchor_text)
        self._current_text = []
        self._in_anchor = False
        self._current_href = None

    def handle_starttag(self, tag, attrs):
        if tag == "a":
            attrs_dict = dict(attrs)
            href = attrs_dict.get("href", "").strip()
            if href and not href.startswith("#") and not href.startswith("mailto:"):
                self._in_anchor = True
                self._current_href = href
                self._current_text = []

    def handle_endtag(self, tag):
        if tag == "a" and self._in_anchor:
            self.links.append((self._current_href, "".join(self._current_text).strip()))
            self._in_anchor = False
            self._current_href = None
            self._current_text = []

    def handle_data(self, data):
        if self._in_anchor:
            self._current_text.append(data)


def extract_links_from_html(html: str) -> list[tuple[str, str]]:
    parser = LinkExtractor()
    parser.feed(html)
    return parser.links


# ─── MD link extractor ───────────────────────────────────────────────────────

def extract_links_from_md(md_text: str) -> list[tuple[str, str]]:
    """Returns list of (href, anchor_text) from markdown [text](url) patterns."""
    return re.findall(r'\[([^\]]+)\]\((https?://[^)]+|/[^)]*)\)', md_text)


def get_md_url(md_text: str) -> str | None:
    """Extract the URL: field from an MD file (bold or plain)."""
    m = re.search(r'\*{0,2}URL:\*{0,2}\s*(https?://\S+)', md_text)
    return m.group(1).rstrip("/") if m else None


# ─── Filtering ───────────────────────────────────────────────────────────────

def is_affiliate(href: str) -> bool:
    try:
        domain = urlparse(href).netloc.lstrip("www.")
        return any(domain == a or domain.endswith("." + a) for a in AFFILIATE_DOMAINS)
    except Exception:
        return False


def is_relative(href: str) -> bool:
    return href.startswith("/") and not href.startswith("//")


def to_relative(href: str, site_domain: str) -> str:
    """Convert absolute same-domain URL to relative path."""
    parsed = urlparse(href)
    if parsed.netloc and site_domain in parsed.netloc:
        return parsed.path.rstrip("/") or "/"
    return href


# ─── Pretty tables ───────────────────────────────────────────────────────────

def col_widths(rows: list[list[str]]) -> list[int]:
    if not rows:
        return []
    return [max(len(str(r[i])) for r in rows) for i in range(len(rows[0]))]


def print_table(headers: list[str], rows: list[list[str]], title: str = ""):
    if title:
        print(f"\n{'─'*80}")
        print(f"  {title}")
        print(f"{'─'*80}")
    all_rows = [headers] + rows
    widths = col_widths(all_rows)
    fmt = "  " + "  │  ".join(f"{{:<{w}}}" for w in widths)
    sep = "  " + "──┼──".join("─" * w for w in widths)
    print(fmt.format(*headers))
    print(sep)
    for row in rows:
        print(fmt.format(*row))


# ─── Step 1: list sites ──────────────────────────────────────────────────────

def list_sites() -> list[tuple[str, str]]:
    """Returns list of (folder_path, site_url)."""
    print("\nConnecting to server and mapping WordPress sites...")
    cmd = f"""
root={WP_ROOT}
wp option get siteurl --path="$root" 2>/dev/null && echo "ROOT_SITE" || true
for d in $root/website_*/; do
  url=$(wp option get siteurl --path="$d" 2>/dev/null) || url='(fail)'
  echo "$d|||$url"
done
"""
    output = ssh(cmd)
    sites = []
    root_url = None
    for line in output.splitlines():
        if "ROOT_SITE" in line:
            continue
        if line.startswith("http") and "|||" not in line:
            root_url = line.strip()
            continue
        if "|||" in line:
            folder, url = line.split("|||", 1)
            folder = folder.strip()
            url = url.strip()
            if url and url != "(fail)" and url.startswith("http"):
                sites.append((folder, url))

    if root_url:
        sites.insert(0, (f"{WP_ROOT}/", root_url))

    return sites


# ─── Step 2: list posts/pages ─────────────────────────────────────────────────

def _fetch_post_type(wp_path: str, site_url: str, post_type: str) -> list[dict]:
    """Single SSH call for one post type, returns parsed list."""
    import json
    domain = urlparse(site_url).netloc.lstrip("www.")
    cmd = (
        f'wp post list --path="{wp_path}" --post_type={post_type} '
        f'--post_status=publish --fields=ID,post_title,post_name '
        f'--format=json 2>/dev/null'
    )
    raw = ssh(cmd).strip()
    if not raw:
        return []
    # Strip any WP-CLI notices/warnings before the JSON array
    json_start = raw.find("[")
    if json_start > 0:
        print(f"  ⚠  WP-CLI noise before JSON ({post_type}): {raw[:json_start][:120]!r}")
        raw = raw[json_start:]
    try:
        rows = json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"  ⚠  JSON parse error ({post_type}): {e}")
        print(f"     Raw (first 300): {raw[:300]!r}")
        return []
    return [
        {
            "id":       str(row.get("ID", "")),
            "title":    row.get("post_title", ""),
            "slug":     row.get("post_name", ""),
            "type":     post_type,
            "wp_path":  wp_path,
            "site_url": site_url,
            "domain":   domain,
        }
        for row in rows
    ]


def list_posts(wp_path: str, site_url: str) -> list[dict]:
    """Returns all published posts + pages as separate SSH calls (avoids JSON concat bug)."""
    posts = _fetch_post_type(wp_path, site_url, "post")
    pages = _fetch_post_type(wp_path, site_url, "page")
    return posts + pages


def get_permalink(post: dict) -> str:
    cmd = f'wp post get {post["id"]} --path="{post["wp_path"]}" --field=guid 2>/dev/null'
    url = ssh(cmd).strip()
    if not url:
        url = post["site_url"].rstrip("/") + "/" + post["slug"] + "/"
    return url


def get_post_content(post: dict) -> str:
    cmd = f'wp post get {post["id"]} --path="{post["wp_path"]}" --field=post_content 2>/dev/null'
    return ssh(cmd)


# ─── Excel slug loader ───────────────────────────────────────────────────────

def _slug_from_cell(value: str) -> str:
    """Extract the final path segment from a cell that may be a full path or bare slug."""
    value = str(value).strip().strip("/")
    return value.split("/")[-1] if "/" in value else value


def load_excel_slugs(site_domain: str) -> dict[str, str]:
    """
    Returns {slug: article_title} from the site's Excel planner.

    Handles two formats:
      Format A (MSM): single 'Master Article List' sheet, 'URL Slug' column has bare slugs
      Format B (Hagia Sofia): per-category sheets (any sheet with an 'Article Title' +
        'URL Slug' column), slug column may contain full paths like /tickets/skip-the-line/
    """
    if not HAS_OPENPYXL:
        return {}

    clean_domain = site_domain.lstrip("www.")
    folder_name  = SITE_TO_INPUT.get(clean_domain, clean_domain.split(".")[0])
    input_dir    = INPUT_BASE / folder_name

    xlsx_files = list(input_dir.glob("*.xlsx"))
    if not xlsx_files:
        return {}

    wb     = openpyxl.load_workbook(xlsx_files[0], read_only=True)
    result = {}

    # ── Format A: single master sheet ────────────────────────────────────────
    if "Master Article List" in wb.sheetnames:
        ws   = wb["Master Article List"]
        rows = list(ws.iter_rows(values_only=True))
        # Real header is row index 1 (row 0 is the sheet title banner)
        for header_idx in range(min(4, len(rows))):
            headers = [str(c).strip() if c else "" for c in rows[header_idx]]
            if "URL Slug" in headers and "Article Title" in headers:
                break
        slug_col  = headers.index("URL Slug")
        title_col = headers.index("Article Title")
        for row in rows[header_idx + 1:]:
            if not row[0]:
                continue
            slug  = _slug_from_cell(row[slug_col])  if row[slug_col]  else ""
            title = str(row[title_col]).strip()      if row[title_col] else ""
            if slug:
                result[slug] = title

    # ── Format B: per-category sheets ────────────────────────────────────────
    else:
        SKIP_SHEETS = {"IA Overview", "Internal Linking Map",
                       "Affiliate URL Builder", "Writing Checklist"}
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            ws   = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            # Find the header row (first row that has 'URL Slug' in it)
            header_idx = None
            for i, row in enumerate(rows):
                if row and any(str(c).strip() == "URL Slug" for c in row if c):
                    header_idx = i
                    break
            if header_idx is None:
                continue
            headers   = [str(c).strip() if c else "" for c in rows[header_idx]]
            slug_col  = headers.index("URL Slug")
            title_col = next(
                (i for i, h in enumerate(headers) if h == "Article Title"), None
            )
            if title_col is None:
                continue
            for row in rows[header_idx + 1:]:
                if not row[0]:
                    continue
                slug  = _slug_from_cell(row[slug_col])  if row[slug_col]  else ""
                title = str(row[title_col]).strip()      if row[title_col] else ""
                if slug:
                    result[slug] = title

    print(f"  Loaded {len(result)} slugs from {xlsx_files[0].name}")
    return result


# ─── Step 3: build MD index ──────────────────────────────────────────────────

def build_md_index(site_domain: str) -> dict[str, tuple[Path, str]]:
    """
    Returns {url_path: (file_path, md_text)} where url_path is like
    '/tickets/ticket-prices' (no trailing slash, no domain).
    """
    # Determine local input folder
    clean_domain = site_domain.lstrip("www.")
    folder_name = SITE_TO_INPUT.get(clean_domain)
    if not folder_name:
        # Try to guess: take second-level of domain
        folder_name = clean_domain.split(".")[0]

    input_dir = INPUT_BASE / folder_name
    if not input_dir.exists():
        print(f"\n  ⚠  Local input folder not found: {input_dir}")
        print(f"     Add an entry to SITE_TO_INPUT in this script to fix.")
        return {}

    index = {}
    for md_file in input_dir.rglob("*.md"):
        text = md_file.read_text(encoding="utf-8", errors="ignore")
        url = get_md_url(text)
        if url:
            path = urlparse(url).path.rstrip("/") or "/"
            index[path] = (md_file, text)

    print(f"\n  Found {len(index)} local MD files in {input_dir.relative_to(SCRIPT_DIR)}")
    return index


# ─── Step 4: audit links for one post ────────────────────────────────────────

def audit_post_links(post: dict, html_content: str, md_index: dict, site_domain: str):
    """
    Returns dict with:
      - permalink
      - link_rows: list of [href, anchor_text, is_relative, in_md, verdict, suggestion]
      - anomalies: list of strings
    """
    all_links = extract_links_from_html(html_content)

    # Filter out affiliates and empty
    links = [
        (href, text) for href, text in all_links
        if href and not is_affiliate(href)
    ]

    # Find matching MD file by post slug / path
    permalink = get_permalink(post)
    post_path = urlparse(permalink).path.rstrip("/") or "/"
    md_entry  = md_index.get(post_path)
    md_links_raw = []
    md_links_by_text = {}
    md_links_by_path = {}

    if md_entry:
        _, md_text = md_entry
        raw = extract_links_from_md(md_text)
        md_links_raw = [(text, href) for text, href in raw if not is_affiliate(href)]
        for text, href in md_links_raw:
            path = urlparse(href).path.rstrip("/") or "/"
            md_links_by_text[text.lower()] = path
            md_links_by_path[path] = text

    link_rows = []
    anomalies = []

    for href, anchor_text in links:
        parsed = urlparse(href)
        is_rel = is_relative(href)
        is_same_domain = site_domain in (parsed.netloc or "")
        href_path = parsed.path.rstrip("/") or "/"

        # Should this be a relative link?
        should_be_relative = is_same_domain or (not parsed.scheme and not parsed.netloc)

        verdict = "✅ OK"
        suggestion = ""

        # Check 1: absolute link that should be relative
        if not is_rel and is_same_domain:
            verdict = "❌ ABS"
            suggestion = f"Change to: {href_path or '/'}"

        # Check 2: does the link target exist in the MD source?
        if md_entry:
            text_key = anchor_text.lower().strip()
            if href_path in md_links_by_path:
                # path matches
                if verdict == "✅ OK":
                    pass  # already good
            elif text_key in md_links_by_text:
                expected_path = md_links_by_text[text_key]
                if href_path != expected_path:
                    if verdict == "✅ OK":
                        verdict = "⚠ WRONG"
                    suggestion += f" MD expects: {expected_path}"
            else:
                # Link not found in MD at all
                if verdict == "✅ OK":
                    verdict = "⚠ NOT IN MD"

        link_rows.append([href, anchor_text, "✓" if is_rel else "✗", verdict, suggestion])

    # Anomaly: posts with no matching MD file
    if not md_entry:
        anomalies.append(f"No local MD file found for path: {post_path}")

    return {
        "permalink": permalink,
        "link_rows": link_rows,
        "anomalies": anomalies,
    }


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 80)
    print("  WordPress Site Link Auditor")
    print("  Usage: python3 link-auditor.py [site-domain] [local-input-folder]")
    print("  Example: python3 link-auditor.py hagiasophia-guide.com hagia-sofia")
    print("=" * 80)

    # CLI args: argv[1] = site domain fragment, argv[2] = local input folder name
    cli_target     = sys.argv[1].lower().strip() if len(sys.argv) > 1 else None
    cli_folder     = sys.argv[2].strip()          if len(sys.argv) > 2 else None

    # 1. List sites
    sites = list_sites()
    if not sites:
        print("No sites found. Check SSH connection.")
        sys.exit(1)

    print("\n  Sites on server:\n")
    for i, (folder, url) in enumerate(sites, 1):
        short_folder = folder.replace(WP_ROOT + "/", "").rstrip("/") or "(root)"
        print(f"  {i:2}.  {url:<45}  [{short_folder}]")

    # 2. Resolve site — CLI arg or interactive prompt
    wp_folder, site_url = None, None
    if cli_target:
        for folder, url in sites:
            if cli_target in url.lower():
                wp_folder, site_url = folder, url
                break
        if not site_url:
            print(f"\n  No site matching '{cli_target}' found on server.")
            sys.exit(1)
        print(f"\n  Auto-selected: {site_url}")
    else:
        print()
        choice = input("  Enter site number to audit: ").strip()
        try:
            idx = int(choice) - 1
            wp_folder, site_url = sites[idx]
        except (ValueError, IndexError):
            print("Invalid choice.")
            sys.exit(1)

    site_domain = urlparse(site_url).netloc.lstrip("www.")

    # Resolve local input folder: CLI arg → prompt
    if cli_folder:
        local_folder = cli_folder
    else:
        available = sorted(p.name for p in INPUT_BASE.iterdir() if p.is_dir())
        print(f"\n  Available local input folders: {', '.join(available)}")
        local_folder = input("  Enter local input folder name for this site: ").strip()

    if not (INPUT_BASE / local_folder).exists():
        print(f"\n  ⚠  Folder not found: {INPUT_BASE / local_folder}")
        sys.exit(1)

    # Register the mapping for all helper functions
    SITE_TO_INPUT[site_domain] = local_folder
    SITE_TO_INPUT[site_domain.lstrip("www.")] = local_folder

    print(f"\n  Auditing: {site_url}  ({wp_folder})")
    print(f"  Local folder: input/{local_folder}")

    # 3. Get all posts/pages
    print("\n  Fetching posts and pages via WP-CLI...")
    posts = list_posts(wp_folder, site_url)
    print(f"  Found {len(posts)} published posts/pages")

    # 4a. Load Excel slugs for cross-check
    excel_slugs = load_excel_slugs(site_domain)

    # ── TABLE 1: all posts/pages with Excel slug check ────────────────────────
    # Build a lookup: excel_slug → title (for "should be" suggestions)
    # Also build MD slug set for cross-reference
    clean_domain  = site_domain.lstrip("www.")
    folder_name   = SITE_TO_INPUT.get(clean_domain, clean_domain.split(".")[0])
    input_dir     = INPUT_BASE / folder_name
    md_slug_set   = set()
    if input_dir.exists():
        for md_file in input_dir.rglob("*.md"):
            text = md_file.read_text(encoding="utf-8", errors="ignore")
            m = re.search(r'\*{0,2}URL:\*{0,2}\s*(https?://\S+)', text)
            if m:
                path = urlparse(m.group(1)).path.strip("/")
                md_slug_set.add(path.split("/")[-1])

    # Silo/category pages that are intentionally absent from Excel
    SILO_SLUGS = {"tickets", "plan-your-visit", "what-to-see", "explore",
                  "getting-there", "home"}

    table1_rows   = []
    slug_problems = []   # posts whose WP slug ≠ Excel slug

    for i, p in enumerate(posts):
        slug = p["slug"]

        if not excel_slugs:
            xl_status = "(no xlsx)"
            should_be = ""
        elif slug in excel_slugs:
            xl_status = "✅ match"
            should_be = ""
        elif slug in SILO_SLUGS or p["type"] == "page":
            xl_status = "— silo page"
            should_be = ""
        else:
            xl_status = "❌ not in Excel"
            # Guess the correct slug: find closest Excel slug whose title
            # shares significant words with the WP title
            best_slug = ""
            best_score = 0
            wp_words = set(re.findall(r'[a-z]+', p["title"].lower()))
            for es, et in excel_slugs.items():
                if es in {r[4] for r in table1_rows}:   # already claimed
                    continue
                ex_words = set(re.findall(r'[a-z]+', et.lower()))
                score = len(wp_words & ex_words)
                if score > best_score:
                    best_score, best_slug = score, es
            should_be = best_slug if best_score >= 3 else "?"
            slug_problems.append((i+1, p, slug, should_be))

        table1_rows.append([str(i+1), p["id"], p["type"], p["title"][:60],
                            slug, xl_status, should_be])

    print_table(
        ["#", "ID", "Type", "Title", "WP Slug", "Excel?", "Should be"],
        table1_rows,
        title="TABLE 1 — All Posts & Pages (Excel slug check)"
    )

    # ── Slug mismatch resolution ──────────────────────────────────────────────
    if excel_slugs:
        wp_slugs     = {p["slug"] for p in posts}
        in_excel_only = {s: t for s, t in excel_slugs.items() if s not in wp_slugs}

        if slug_problems or in_excel_only:
            print(f"\n{'═'*80}")
            print("  SLUG MISMATCH RESOLUTION")
            print(f"{'═'*80}")
            print("  Fix WP slugs before proceeding to link audit.\n")

            # Map Excel-only slugs by title similarity to unmatched WP posts
            unmatched_posts = [p for p in posts
                               if p["slug"] not in excel_slugs and p["type"] != "page"
                               and p["slug"] not in SILO_SLUGS]

            if in_excel_only:
                print(f"  Excel slugs with no matching WP post ({len(in_excel_only)}):\n")
                for s, t in in_excel_only.items():
                    print(f"    Excel:  {s:<45}  \"{t}\"")

            if slug_problems:
                print(f"\n  WP posts not matching any Excel slug ({len(slug_problems)}):\n")
                for num, post, wp_slug, suggested in slug_problems:
                    print(f"    #{num:>2}  WP slug : {wp_slug}")
                    print(f"          Title   : {post['title']}")
                    print(f"          Best match: {suggested or '?'}")
                    print()

                print("  For each mismatched post, type the correct slug, Enter to accept")
                print("  the suggestion, 's' to skip, or 'q' to quit.\n")

                slug_renames = {}   # post_id → (old_slug, new_slug)
                for num, post, wp_slug, suggested in slug_problems:
                    prompt_default = f" [{suggested}]" if suggested and suggested != "?" else ""
                    answer = input(
                        f"  #{num} \"{post['title'][:55]}\"  →  slug{prompt_default}: "
                    ).strip()
                    if answer == "q":
                        break
                    if answer in ("s", ""):
                        # Empty + no suggestion → skip; empty + suggestion → accept suggestion
                        if answer == "" and suggested and suggested != "?":
                            answer = suggested
                        else:
                            continue
                    if answer and answer != wp_slug:
                        slug_renames[post["id"]] = (wp_slug, answer)

                if slug_renames:
                    print(f"\n  Applying {len(slug_renames)} slug rename(s) via WP-CLI...")
                    for post_id, (old_slug, new_slug) in slug_renames.items():
                        out = ssh(
                            f'wp post update {post_id} --post_name="{new_slug}" '
                            f'--path="{wp_folder}" 2>/dev/null'
                        )
                        # Also update post URL in the posts list for Table 2
                        for p in posts:
                            if p["id"] == post_id:
                                p["slug"] = new_slug
                        print(f"    {old_slug}  →  {new_slug}  ({out.strip() or 'ok'})")
                    print()

    # 4b. Build local MD index
    md_index = build_md_index(site_domain)

    # 5. Per-post interactive link audit
    print(f"\n\n{'═'*80}")
    print("  TABLE 2 — Per-Post Link Audit (non-affiliate links only)")
    print(f"{'═'*80}")
    print("  For each post: review links → confirm fixes → changes applied via WP-CLI")
    print("  Commands at each post:  [y] apply fixes   [n] skip   [q] quit\n")

    all_anomalies = []
    fixed_count   = 0

    for i, post in enumerate(posts):
        print(f"\n{'─'*80}")
        print(f"  [{i+1}/{len(posts)}]  {post['title'][:75]}")

        html_content = get_post_content(post)
        if not html_content.strip():
            print("  (no content returned — skipping)")
            continue

        result   = audit_post_links(post, html_content, md_index, site_domain)
        permalink = result["permalink"]
        link_rows = result["link_rows"]

        print(f"  URL: {permalink}\n")

        # Separate issues from OK rows
        issue_rows = [r for r in link_rows if not r[3].startswith("✅")]
        ok_rows    = [r for r in link_rows if r[3].startswith("✅")]

        if not issue_rows:
            print("  All links OK — nothing to fix.")
            if result["anomalies"]:
                for a in result["anomalies"]:
                    print(f"  ⚠  {a}")
                    all_anomalies.append(f"[{post['title'][:50]}] {a}")
            continue

        # Show only problematic links
        print_table(
            ["#", "Current href", "Anchor Text", "Rel?", "Status", "Suggested fix"],
            [[str(j+1)] + r for j, r in enumerate(issue_rows)],
            title=f"Issues found ({len(issue_rows)} link(s) need attention, {len(ok_rows)} OK)"
        )

        if result["anomalies"]:
            for a in result["anomalies"]:
                print(f"\n  ⚠  {a}")
                all_anomalies.append(f"[{post['title'][:50]}] {a}")

        # Build the set of fixes to apply
        fixes = []  # list of (old_href, new_href)
        for href, anchor_text, is_rel, verdict, suggestion in issue_rows:
            if "❌ ABS" in verdict:
                # Extract the suggested relative path
                m = re.search(r'Change to:\s*(\S+)', suggestion)
                if m:
                    fixes.append((href, m.group(1)))
            elif "⚠ WRONG" in verdict:
                m = re.search(r'MD expects:\s*(\S+)', suggestion)
                if m:
                    fixes.append((href, m.group(1)))
            # NOT IN MD — no automatic fix, just flagged

        if not fixes:
            print("\n  (no auto-fixable issues — manual review needed)")
            continue

        print(f"\n  Proposed changes ({len(fixes)}):")
        for old, new in fixes:
            print(f"    {old}")
            print(f"    → {new}\n")

        answer = input("  Apply these fixes? [y/n/q]: ").strip().lower()
        if answer == "q":
            print("  Quitting.")
            break
        if answer != "y":
            print("  Skipped.")
            continue

        # Apply fixes: replace hrefs in post content and update via WP-CLI
        updated_content = html_content
        applied = []
        for old_href, new_href in fixes:
            # Replace href="<old>" with href="<new>" (quoted, various quote styles)
            old_escaped = re.escape(old_href)
            updated_content, n1 = re.subn(
                rf'href=["\']({old_escaped})["\']',
                lambda m, new=new_href: m.group(0).replace(m.group(1), new),
                updated_content
            )
            if n1:
                applied.append((old_href, new_href, n1))

        if not applied:
            print("  No replacements made (href pattern not found in raw content).")
            continue

        # Write updated content back via WP-CLI using a temp file approach
        import tempfile
        with tempfile.NamedTemporaryFile(mode="w", suffix=".html", delete=False) as tf:
            tf.write(updated_content)
            tmp_path = tf.name

        remote_tmp = f"/tmp/wp_content_{post['id']}.html"
        # Upload temp file to server, then use WP-CLI to update
        upload_cmd = ["scp", "-i", SSH_KEY, "-o", "StrictHostKeyChecking=no",
                      tmp_path, f"{SSH_USER}@{SSH_HOST}:{remote_tmp}"]
        subprocess.run(upload_cmd, capture_output=True)
        os.unlink(tmp_path)

        update_cmd = (
            f'content=$(cat {remote_tmp}); '
            f'wp post update {post["id"]} --post_content="$content" '
            f'--path="{post["wp_path"]}" 2>/dev/null; '
            f'rm -f {remote_tmp}'
        )
        out = ssh(update_cmd)
        print(f"\n  WP-CLI: {out or '(ok)'}")

        for old_href, new_href, count in applied:
            print(f"  ✅ Replaced {count}x  {old_href}  →  {new_href}")
            fixed_count += 1

    # 6. Summary
    print(f"\n\n{'═'*80}")
    print("  SESSION SUMMARY")
    print(f"{'═'*80}")
    print(f"  Posts/pages audited : {len(posts)}")
    print(f"  Links fixed         : {fixed_count}")

    if all_anomalies:
        print(f"\n  Anomalies (no auto-fix):")
        for a in all_anomalies:
            print(f"    ⚠  {a}")
    else:
        print("\n  No structural anomalies detected.")

    print(f"\n  Done.")


if __name__ == "__main__":
    main()
