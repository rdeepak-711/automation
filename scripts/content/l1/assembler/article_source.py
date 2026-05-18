#!/usr/bin/env python3
"""
article_source.py — Load plan-your-visit articles from xlsx + l1-config.

Returns list of article dicts for the Plan Your Visit silo:
  {idx, title, url_slug, url, tags, subgroup, meta_desc}

Tags come from l1-config._card_tags (or empty list).
Subgroup comes from l1-config._subgroups (or None).
meta_desc is read from the MD file intro if available.
"""

import json
import re
from pathlib import Path

PLAN_VISIT_KEYWORDS = {"plan", "visit", "planning", "how to"}


def _is_what_to_see_category(category: str) -> bool:
    cat = category.lower()
    return "see" in cat and "plan" not in cat and "ticket" not in cat and "tour" not in cat


def _load_article_metas(site_slug: str, repo_root) -> dict:
    """Load consolidated article-metas.json if available. Returns {} if missing."""
    p = Path(repo_root) / "output" / site_slug / "article-metas.json"
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _is_plan_your_visit_category(category: str) -> bool:
    cat = category.lower()
    return "plan" in cat or "visit" in cat


def _read_md_meta_desc(md_dir: Path, url_slug: str) -> str:
    """Return the Meta Description from the MD file for this slug, or ''."""
    if not md_dir.is_dir():
        return ""
    for md_file in md_dir.glob("*.md"):
        stem = md_file.stem
        # Filename format: article{N}_{slug}.md
        if stem.endswith(url_slug.replace("-", "-")):
            slug_part = re.sub(r"^article\d+_", "", stem)
            if slug_part == url_slug.replace("-", "-"):
                text = md_file.read_text(encoding="utf-8")
                m = re.search(r"\*\*Meta Description:\*\*\s*(.+)", text)
                if m:
                    return m.group(1).strip()
    return ""


def _parse_xlsx(xlsx_path: Path, sheet_name: str = "IA — All Articles") -> list:
    """Parse xlsx and return all rows as {title, category, url, notes}."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)

    SILO_SHEETS = ["Tickets & Tours", "Plan Your Visit", "What To See", "What to See"]
    has_silo_sheets = any(s in wb.sheetnames for s in SILO_SHEETS)

    rows = []
    if has_silo_sheets:
        for sheet_name in wb.sheetnames:
            if not any(s.lower() in sheet_name.lower() for s in ["ticket", "plan", "see"]):
                continue
            ws = wb[sheet_name]
            category = sheet_name
            for row in ws.iter_rows(min_row=2):
                vals = [c.value for c in row]
                if not vals[0] or not vals[1]:
                    continue
                first = str(vals[0]).strip().upper()
                if first in ("PILLAR", "CLUSTER", "#", ""):
                    continue
                if not first[0:1].isalnum():
                    continue
                url = str(vals[3]).strip() if len(vals) > 3 and vals[3] else ""
                rows.append({
                    "title": str(vals[1]).strip(),
                    "category": category,
                    "url": url,
                })
    else:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
        current_cat = None
        is_standard = None
        for row in ws.iter_rows(min_row=2):
            vals = [c.value for c in row]
            if all(v is None for v in vals):
                continue
            col0 = str(vals[0]).strip() if vals[0] else ""
            if col0 and "\n" in col0:
                current_cat = col0.split("\n")[0].strip()
                is_standard = True
                # First article may be on the same row as the category header
                if not (vals[1] and vals[2]):
                    continue
            if is_standard is None:
                is_standard = bool(
                    vals[1]
                    and str(vals[1]).strip()
                    and len(str(vals[1]).strip()) <= 4
                    and str(vals[1]).strip()[0] in "PTWA"
                )
            if is_standard:
                cat_id, title, url_slug = vals[1], vals[2], vals[3]
                if not cat_id or not title or not current_cat:
                    continue
                rows.append({
                    "title": str(title).strip(),
                    "category": current_cat,
                    "url": str(url_slug or "").strip(),
                })
            else:
                if not vals[0] or not vals[1]:
                    continue
                full_url = str(vals[5]).strip() if len(vals) > 5 and vals[5] else ""
                rows.append({
                    "title": str(vals[1]).strip(),
                    "category": str(vals[2] or "").strip(),
                    "url": full_url,
                })
    return rows


def _slug_from_url(url: str) -> str:
    """Extract the final path segment as the slug."""
    url = url.rstrip("/")
    if "/" in url:
        return url.rsplit("/", 1)[-1]
    return url


def _is_tickets_category(category: str) -> bool:
    cat = category.lower()
    return "ticket" in cat or "tour" in cat


def _read_tickets_md(tickets_path: Path) -> list[dict]:
    """Parse tickets.md — pipe-delimited: Tid | Title | URL | token | /slug/path/"""
    tickets = []
    if not tickets_path.exists():
        return tickets
    for line in tickets_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if not line[0].upper().startswith("T") or "|" not in line:
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 5:
            continue
        tid = parts[0]
        title = re.sub(r"\s+[Pp]\d{4,}$", "", parts[1]).strip()  # strip Tiqets product ID suffix
        affiliate_url = parts[2]
        col5 = parts[4]
        article_slug = col5.rstrip("/").rsplit("/", 1)[-1] if "/" in col5.rstrip("/") else col5
        article_url = col5.rstrip("/") + "/" if col5 else f"/tickets/{article_slug}/"
        tickets.append({
            "tid": tid,
            "title": title,
            "affiliate_url": affiliate_url,
            "article_slug": article_slug,
            "article_url": article_url,
        })
    return tickets


def _load_sidecar(output_dir: Path, slug: str) -> dict:
    """Load .meta.json sidecar for an article slug. Returns {} if missing."""
    for candidate in output_dir.rglob(f"{slug}.meta.json"):
        try:
            return json.loads(candidate.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _classify_articles_claude(
    articles: list[dict],
    cache: dict,
    model: str = "claude-haiku-4-5-20251001",
) -> dict[str, str]:
    """
    Use Claude to classify each editorial article as 'ticket' or 'info'.
    ticket = booking/experience-focused (helps visitor choose/book a specific experience)
    info   = planning/comparison-focused (general guidance, FAQs, logistics)
    Returns {url_slug: 'ticket'|'info'}. Cached under '_article_types'.
    """
    import subprocess

    CACHE_KEY = "_article_types"
    if CACHE_KEY in cache:
        return cache[CACHE_KEY]

    block = "\n".join(
        f'  {a["url_slug"]}: "{a["title"]}" — tags: {a.get("tags", [])}'
        for a in articles
    )
    prompt = f"""Classify each travel article as either 'ticket' or 'info'.

ticket = article helps visitor choose or book a specific experience/ticket/tour
info   = article gives general planning guidance, FAQs, logistics, comparisons

Articles:
{block}

Return ONLY a JSON object: {{"slug": "ticket", "slug2": "info", ...}}
Use exact slugs as shown. No explanation."""

    try:
        r = subprocess.run(
            ["claude", "--print", "--model", model],
            input=prompt, capture_output=True, text=True, timeout=60,
        )
        raw = r.stdout.strip() if r.returncode == 0 else ""
    except (FileNotFoundError, subprocess.TimeoutExpired):
        raw = ""

    result = {}
    if raw:
        try:
            result = json.loads(raw)
        except json.JSONDecodeError:
            m = re.search(r"\{[^{}]+\}", raw, re.DOTALL)
            if m:
                try:
                    result = json.loads(m.group(0))
                except json.JSONDecodeError:
                    pass

    if not result:
        # Fallback: all articles classified as ticket
        result = {a["url_slug"]: "ticket" for a in articles}

    cache[CACHE_KEY] = result
    return result


def _match_tickets_to_articles_claude(
    top_tickets: list[dict],
    editorial_articles: list[dict],
    cache: dict,
    model: str = "claude-haiku-4-5-20251001",
) -> dict[str, str]:
    """
    Use Claude to match each top ticket (GYG product) to the best editorial article.
    Returns {ticket_tid: article_url_slug} mapping.
    Caches result under '_featured_match' in cache dict (l1-config.json).
    """
    import subprocess

    CACHE_KEY = "_featured_match"
    if CACHE_KEY in cache:
        return cache[CACHE_KEY]

    tickets_block = "\n".join(
        f'  {t["tid"]}: "{t["title"]}"' for t in top_tickets
    )
    articles_block = "\n".join(
        f'  {a["url_slug"]}: "{a["title"]}" — {a.get("description", "")[:120]}'
        for a in editorial_articles
    )
    prompt = f"""Match each ticket product to the single best editorial article guide.

Ticket products (GYG/Tiqets):
{tickets_block}

Editorial article guides:
{articles_block}

Return ONLY a JSON object mapping each ticket TID to the best article slug:
{{"T1": "slug-of-best-article", "T2": "slug-of-best-article", ...}}
Each TID maps to exactly one slug. Use the slug exactly as shown. No explanation."""

    try:
        r = subprocess.run(
            ["claude", "--print", "--model", model],
            input=prompt, capture_output=True, text=True, timeout=60,
        )
        raw = r.stdout.strip() if r.returncode == 0 else ""
    except (FileNotFoundError, subprocess.TimeoutExpired):
        raw = ""

    result = {}
    if raw:
        try:
            result = json.loads(raw)
        except json.JSONDecodeError:
            m = re.search(r"\{[^{}]+\}", raw, re.DOTALL)
            if m:
                try:
                    result = json.loads(m.group(0))
                except json.JSONDecodeError:
                    pass

    if not result:
        # Fallback: sequential positional match
        for i, t in enumerate(top_tickets):
            if i < len(editorial_articles):
                result[t["tid"]] = editorial_articles[i]["url_slug"]

    cache[CACHE_KEY] = result
    return result


def _match_articles_to_affiliate_claude(
    editorial_tickets: list[dict],
    affiliate_tickets: list[dict],
    cache: dict,
    model: str = "claude-haiku-4-5-20251001",
) -> dict[str, str]:
    """
    Match each editorial ticket article to the best affiliate ticket (GYG/Tiqets product).
    Returns {article_url_slug: affiliate_tid} mapping.
    Cached under '_editorial_affiliate_map'.
    """
    import subprocess

    CACHE_KEY = "_editorial_affiliate_map"
    if CACHE_KEY in cache:
        return cache[CACHE_KEY]

    articles_block = "\n".join(
        f'  {a["url_slug"]}: "{a["title"]}"' for a in editorial_tickets
    )
    tickets_block = "\n".join(
        f'  {t["tid"]}: "{t["title"]}"' for t in affiliate_tickets
    )
    prompt = f"""Match each editorial article guide to the single best affiliate ticket product.
Each article is about a type of experience — pick the affiliate ticket that best fits.
Multiple articles may map to the same ticket. Every article must map to exactly one ticket.

Editorial articles:
{articles_block}

Affiliate ticket products:
{tickets_block}

Return ONLY a JSON object: {{article_slug: affiliate_tid, ...}}
Use slugs and TIDs exactly as shown. No explanation."""

    try:
        r = subprocess.run(
            ["claude", "--print", "--model", model],
            input=prompt, capture_output=True, text=True, timeout=60,
        )
        raw = r.stdout.strip() if r.returncode == 0 else ""
    except (FileNotFoundError, subprocess.TimeoutExpired):
        raw = ""

    result: dict[str, str] = {}
    if raw:
        try:
            result = json.loads(raw)
        except json.JSONDecodeError:
            m = re.search(r"\{[^{}]+\}", raw, re.DOTALL)
            if m:
                try:
                    result = json.loads(m.group(0))
                except json.JSONDecodeError:
                    pass

    if not result:
        # Fallback: map each article to the first affiliate ticket
        for a in editorial_tickets:
            if affiliate_tickets:
                result[a["url_slug"]] = affiliate_tickets[0]["tid"]

    cache[CACHE_KEY] = result
    return result


def load_plan_your_visit_articles(site_slug: str, repo_root: str | None = None) -> list[dict]:
    """
    Returns articles for the Plan Your Visit silo.

    Each dict:
      {idx, title, url_slug, url, tags, subgroup, meta_desc}
    """
    if repo_root is None:
        repo_root = Path(__file__).parent.parent.parent.parent
    else:
        repo_root = Path(repo_root)

    xlsx_candidates = list((repo_root / "input" / site_slug).glob("*.xlsx"))
    if not xlsx_candidates:
        raise FileNotFoundError(f"No xlsx found in input/{site_slug}/")
    xlsx_path = xlsx_candidates[0]

    config_path = repo_root / "input" / site_slug / "l1-config.json"
    cfg = {}
    if config_path.exists():
        cfg = json.loads(config_path.read_text(encoding="utf-8"))

    card_tags = cfg.get("_card_tags", {})
    subgroups = cfg.get("_subgroups", {})

    # Determine xlsx sheet name from config
    sheet_name = cfg.get("sheet_name", "IA — All Articles")
    all_rows = _parse_xlsx(xlsx_path, sheet_name)

    # Filter plan-your-visit articles
    pyv_rows = [r for r in all_rows if _is_plan_your_visit_category(r["category"])]

    # MD directory for reading meta descriptions
    md_dir = repo_root / "input" / site_slug / "plan-your-visit"

    result = []
    for idx, row in enumerate(pyv_rows):
        title = row["title"]
        raw_url = row["url"]
        url_slug = _slug_from_url(raw_url)
        # Build canonical URL
        if raw_url.startswith("http"):
            # Strip domain, keep path
            m = re.search(r"https?://[^/]+(/.+)", raw_url)
            url = m.group(1).rstrip("/") + "/" if m else f"/plan-your-visit/{url_slug}/"
        elif raw_url.startswith("/"):
            url = raw_url.rstrip("/") + "/"
        else:
            url = f"/plan-your-visit/{url_slug}/"

        # Normalise protocol-relative URLs (//path → /path)
        url = re.sub(r"^//", "/", url)

        tags = card_tags.get(title, [])
        subgroup = subgroups.get(title)
        meta_desc = _read_md_meta_desc(md_dir, url_slug)

        result.append({
            "idx": idx,
            "title": title,
            "url_slug": url_slug,
            "url": url,
            "tags": tags,
            "subgroup": subgroup,
            "meta_desc": meta_desc,
        })

    return result


def get_faq_url(site_slug: str, repo_root: str | None = None) -> str:
    """Return the FAQ page URL for this site from the xlsx IA sheet.

    Looks for a plan-your-visit article whose url contains 'faq'.
    Falls back to '/faqs/' if not found or xlsx is unavailable.
    """
    try:
        articles = load_plan_your_visit_articles(site_slug, repo_root)
        for a in articles:
            if "faq" in a.get("url", "").lower() or "faq" in a.get("url_slug", "").lower():
                return a["url"]
    except Exception:
        pass
    return "/faqs/"


def load_what_to_see_articles(site_slug: str, repo_root: str | None = None) -> list[dict]:
    """
    Returns articles for the What to See silo.

    Each dict:
      {idx, title, url_slug, url, tags, card_title, description, subgroup, meta_desc}

    card_title and description merged from article-metas.json.
    """
    if repo_root is None:
        repo_root = Path(__file__).parent.parent.parent.parent
    else:
        repo_root = Path(repo_root)

    xlsx_candidates = list((repo_root / "input" / site_slug).glob("*.xlsx"))
    if not xlsx_candidates:
        raise FileNotFoundError(f"No xlsx found in input/{site_slug}/")
    xlsx_path = xlsx_candidates[0]

    config_path = repo_root / "input" / site_slug / "l1-config.json"
    cfg = {}
    if config_path.exists():
        cfg = json.loads(config_path.read_text(encoding="utf-8"))

    card_tags = cfg.get("_card_tags", {})
    subgroups = cfg.get("_subgroups", {})
    sheet_name = cfg.get("sheet_name", "IA — All Articles")
    all_rows = _parse_xlsx(xlsx_path, sheet_name)

    wts_rows = [r for r in all_rows if _is_what_to_see_category(r["category"])]

    md_dir = repo_root / "input" / site_slug / "what-to-see"
    metas = _load_article_metas(site_slug, repo_root)

    result = []
    for idx, row in enumerate(wts_rows):
        title = row["title"]
        raw_url = row["url"]
        url_slug = _slug_from_url(raw_url)
        if raw_url.startswith("http"):
            m = re.search(r"https?://[^/]+(/.+)", raw_url)
            url = m.group(1).rstrip("/") + "/" if m else f"/what-to-see/{url_slug}/"
        elif raw_url.startswith("/"):
            url = raw_url.rstrip("/") + "/"
        else:
            url = f"/what-to-see/{url_slug}/"
        url = re.sub(r"^//", "/", url)

        m_data = metas.get(url_slug, {})
        tags = m_data.get("tags") or card_tags.get(title, [])
        card_title = m_data.get("card_title", "")
        meta_desc = _read_md_meta_desc(md_dir, url_slug)
        description = m_data.get("description") or meta_desc

        result.append({
            "idx": idx,
            "title": title,
            "url_slug": url_slug,
            "url": url,
            "tags": tags,
            "card_title": card_title,
            "description": description,
            "subgroup": subgroups.get(title),
            "meta_desc": meta_desc,
        })

    return result


def load_tickets_tours_articles(site_slug: str, repo_root: str | None = None, force: bool = False) -> dict:
    """
    Returns bundle dict for the Tickets & Tours silo.

    Featured cards: TOP_TICKET_INDEXES → tickets.md entries → keyword-matched to
      editorial L2 articles. Each featured card carries both the article data
      (title, description, tags, url) and the affiliate_url from tickets.md.

    Remaining 22 editorial articles split by tag type:
      tickets      — booking/experience-type articles
      informational — planning/info articles

    affiliate_tickets — all 53 tickets.md rows, for comparison table only.

    Bundle keys:
      featured, tickets, informational,
      tickets_by_tid (editorial, synthetic E-TIDs),
      affiliate_tickets, affiliate_tickets_by_tid,
      tickets_by_slug (tickets.md slug → row)
    """
    if repo_root is None:
        repo_root = Path(__file__).parent.parent.parent.parent
    else:
        repo_root = Path(repo_root)

    xlsx_candidates = list((repo_root / "input" / site_slug).glob("*.xlsx"))
    if not xlsx_candidates:
        raise FileNotFoundError(f"No xlsx found in input/{site_slug}/")
    xlsx_path = xlsx_candidates[0]

    config_path = repo_root / "input" / site_slug / "l1-config.json"
    cfg = {}
    if config_path.exists():
        cfg = json.loads(config_path.read_text(encoding="utf-8"))

    card_tags = cfg.get("_card_tags", {})
    subgroups = cfg.get("_subgroups", {})

    # Load env for TOP_TICKET_INDEXES
    env = {}
    for env_path in [repo_root / ".env", repo_root / "input" / site_slug / ".env"]:
        if env_path.exists():
            for line in env_path.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, _, v = line.partition("=")
                env[k.strip()] = v.strip().strip('"').strip("'")

    raw_idx = env.get("TOP_TICKET_INDEXES", "1,2")
    try:
        top_indexes = [int(x.strip()) for x in raw_idx.split(",") if x.strip()]
    except ValueError:
        top_indexes = [1, 2]

    # ── tickets.md → affiliate_tickets (GYG/Tiqets products, table use only) ──
    tickets_md_rows = _read_tickets_md(repo_root / "input" / site_slug / "tickets.md")
    affiliate_tickets = []
    _seen_aff_urls: set[str] = set()
    for trow in tickets_md_rows:
        aff_url = trow["affiliate_url"]
        if aff_url in _seen_aff_urls:
            continue  # deduplicate exact-URL duplicates from tickets.md
        _seen_aff_urls.add(aff_url)
        affiliate_tickets.append({
            "tid": trow["tid"],
            "title": trow["title"],
            "affiliate_url": aff_url,
            "article_slug": trow["article_slug"],
            "article_url": trow["article_url"],
            "tags": [],
            "description": "",
        })
    affiliate_tickets_by_tid = {t["tid"]: t for t in affiliate_tickets}
    tickets_by_slug_md = {t["article_slug"]: t for t in affiliate_tickets}

    # Top tickets.md entries (by TOP_TICKET_INDEXES, 1-based)
    top_affiliate = []
    for idx in top_indexes:
        if 1 <= idx <= len(affiliate_tickets):
            top_affiliate.append(affiliate_tickets[idx - 1])

    # ── xlsx T&T rows → editorial articles (L2 HTML must exist) ──────────────
    sheet_name = cfg.get("sheet_name", "IA — All Articles")
    all_rows = _parse_xlsx(xlsx_path, sheet_name)
    tt_rows = [r for r in all_rows if _is_tickets_category(r["category"])]

    l2_dir = repo_root / "output" / site_slug / "l2-articles" / "tickets-tours"
    metas = _load_article_metas(site_slug, repo_root)
    info_md_dir = repo_root / "input" / site_slug / "tickets-tours"

    all_editorial: list[dict] = []
    for i, row in enumerate(tt_rows):
        slug = _slug_from_url(row["url"])
        if not slug:
            continue
        if not (l2_dir / f"{slug}.html").exists():
            continue  # skip articles without L2 HTML
        raw_url = row["url"]
        if raw_url.startswith("http"):
            m = re.search(r"https?://[^/]+(/.+)", raw_url)
            url = m.group(1).rstrip("/") + "/" if m else f"/tickets/{slug}/"
        elif raw_url.startswith("/"):
            url = raw_url.rstrip("/") + "/"
        else:
            url = f"/tickets/{slug}/"
        url = re.sub(r"^//", "/", url)
        title = row["title"]
        m_data = metas.get(slug, {})
        tags = m_data.get("tags") or card_tags.get(title, [])
        description = m_data.get("description") or _read_md_meta_desc(info_md_dir, slug)
        all_editorial.append({
            "tid": f"E{i + 1}",   # synthetic TID for card rendering machinery
            "title": title,
            "card_title": m_data.get("card_title", ""),
            "url_slug": slug,
            "url": url,
            "article_url": url,
            "affiliate_url": "",  # filled in for featured articles below
            "article_slug": slug,
            "tags": tags,
            "description": description,
            "subgroup": subgroups.get(title),
            "is_editorial": True,
            "ticket_url": m_data.get("ticket_url", ""),
        })

    # ── Build featured: Claude matches each top ticket → best editorial article ─
    if force:
        cfg.pop("_featured_match", None)
    editorial_by_slug = {a["url_slug"]: a for a in all_editorial}
    had_match_cache = "_featured_match" in cfg
    slug_map = _match_tickets_to_articles_claude(top_affiliate, all_editorial, cfg)
    if not had_match_cache and config_path.exists():
        # Persist the newly computed match so next run is instant
        config_path.write_text(
            json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8"
        )
    featured: list[dict] = []
    featured_slugs: set[str] = set()
    for aff_ticket in top_affiliate:
        matched_slug = slug_map.get(aff_ticket["tid"])
        article = editorial_by_slug.get(matched_slug) if matched_slug else None
        if article is None:
            # Fallback: first unmatched editorial article
            article = next((a for a in all_editorial if a["url_slug"] not in featured_slugs), None)
        if article and article["url_slug"] not in featured_slugs:
            card = dict(article)
            card["affiliate_url"] = aff_ticket["affiliate_url"]  # attach GYG/Tiqets URL
            featured.append(card)
            featured_slugs.add(article["url_slug"])

    # ── Remaining editorial articles → linked (has ticket_url) vs unlinked ──
    # linked  = article has a specific affiliate booking URL (ticket_url from metas)
    #           → rendered in ticket-group sections ABOVE the comparison table
    # unlinked = no affiliate link → rendered in guide sections BELOW table + how-to-pick
    remaining = [a for a in all_editorial if a["url_slug"] not in featured_slugs]
    ticket_list = [a for a in remaining if a.get("ticket_url")]
    info_list = [a for a in remaining if not a.get("ticket_url")]

    # Set affiliate_url on linked articles.
    # For articles without ticket_url, fall back to Claude match against affiliate_tickets.
    needs_match = [a for a in ticket_list if not a.get("ticket_url")]
    if needs_match:
        if force:
            cfg.pop("_editorial_affiliate_map", None)
        editorial_affiliate_map = _match_articles_to_affiliate_claude(needs_match, affiliate_tickets, cfg)
        if config_path.exists():
            config_path.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")
        for a in needs_match:
            matched_tid = editorial_affiliate_map.get(a["url_slug"])
            if matched_tid and matched_tid in affiliate_tickets_by_tid:
                a["affiliate_url"] = affiliate_tickets_by_tid[matched_tid]["affiliate_url"]

    for a in ticket_list:
        if a.get("ticket_url"):
            a["affiliate_url"] = a["ticket_url"]

    # tickets_by_tid: editorial articles keyed by synthetic E-TID (for _render_ticket_groups)
    tickets_by_tid = {a["tid"]: a for a in all_editorial}

    return {
        "tickets": ticket_list,
        "informational": info_list,
        "featured": featured,
        "tickets_by_tid": tickets_by_tid,
        "tickets_by_slug": tickets_by_slug_md,
        "affiliate_tickets": affiliate_tickets,
        "affiliate_tickets_by_tid": affiliate_tickets_by_tid,
    }
