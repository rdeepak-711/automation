#!/usr/bin/env python3
"""
standardize-xlsx.py

One-time converter that normalises all site XLSX files to the standard 7-column format:
  #  |  Category  |  Article Title  |  URL Slug  |  Target Query  |  Affiliate  |  Priority

Usage:
    python3 scripts/content/standardize-xlsx.py
"""

import glob
import os
import shutil
import sys
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

HEADER = ["#", "Category", "Article Title", "URL Slug", "Target Query", "Affiliate", "Priority"]
HEADER_BG = "4472C4"
HEADER_FG = "FFFFFF"

# ─── Category display-name → (L1 path, dir name, id prefix) ──────────────────
CATEGORY_MAP = {
    # MSM / generic names
    "tickets & tours":    ("/tickets/",           "tickets-tours",    "T"),
    "plan your visit":    ("/plan-your-visit/",    "plan-your-visit",  "P"),
    "what to see & do":   ("/what-to-see/",        "what-to-see",      "W"),
    "what to see":        ("/what-to-see/",        "what-to-see",      "W"),
    "what to see and do": ("/what-to-see/",        "what-to-see",      "W"),
    # blue-mosque uses "What To See"
    "what to see":        ("/what-to-see/",        "what-to-see",      "W"),
}

# Affiliate text → short code
AFFILIATE_CLEAN = {
    "tiqets (main ticket)": "Tiqets",
    "tiqets":               "Tiqets",
    "getyourguide":         "GYG",
    "gyg":                  "GYG",
    "viator":               "Viator",
}

PRIORITY_CLEAN = {
    "★★★ high": "High",
    "★★ medium": "Medium",
    "★ low": "Low",
    "high": "High",
    "medium": "Medium",
    "low": "Low",
}


# ─── Format detection ─────────────────────────────────────────────────────────

def detect_format(header_row):
    """Return 'A', 'B', or 'C' (or raise ValueError)."""
    h = [str(c).strip() if c else "" for c in header_row]

    # Format B: 14 cols with 'Full URL' col (check BEFORE Format A)
    if len(h) >= 5 and h[4] == "Full URL":
        return "B"

    # Format A: standard 7 cols — URL Slug col contains full relative paths
    # (no Full URL column present)
    if len(h) >= 7 and h[3] == "URL Slug":
        return "A"

    # Format C: 6 cols — #, Article Title, Category, Category Slug, Article Slug, Full URL
    if len(h) >= 6 and h[2] == "Category" and h[3] == "Category Slug":
        return "C"

    raise ValueError(f"Unknown header format: {h[:7]}")


# ─── Converters ──────────────────────────────────────────────────────────────

def rows_from_ws(ws):
    """Return all rows (including header) as lists of raw values."""
    return [list(row) for row in ws.iter_rows(values_only=True)]


def find_header_row_index(rows):
    """Find the index of the true header row (contains '#' as first non-None)."""
    for i, row in enumerate(rows):
        vals = [str(c).strip() if c else "" for c in row]
        if vals[0] == "#" or (vals[0] == "" and vals[1] == "Category"):
            return i
    raise ValueError("Could not locate header row")


def clean_url_slug(slug):
    """Ensure slug has leading and trailing slash."""
    if not slug:
        return "/"
    slug = str(slug).strip()
    if not slug.startswith("/"):
        slug = "/" + slug
    if not slug.endswith("/"):
        slug = slug + "/"
    return slug


def extract_path_from_url(url):
    """Turn https://domain.com/tickets/foo/ into /tickets/foo/."""
    if not url:
        return "/"
    url = str(url).strip()
    # Remove protocol + domain
    if "://" in url:
        url = url.split("://", 1)[1]
    if "/" in url:
        url = url[url.index("/"):]
    return clean_url_slug(url)


def clean_affiliate(raw):
    if not raw:
        return "—"
    raw_lower = str(raw).strip().lower()
    return AFFILIATE_CLEAN.get(raw_lower, "—")


def clean_priority(raw):
    if not raw:
        return "High"
    raw_lower = str(raw).strip().lower()
    return PRIORITY_CLEAN.get(raw_lower, str(raw).strip())


def normalize_cat_key(name):
    return str(name).strip().lower()


def convert_format_b(rows):
    """MSM 14-col → standard 7-col rows."""
    # Find header
    hi = find_header_row_index(rows)
    data_rows = rows[hi + 1:]

    # Group by category, preserving order
    cat_order = []
    cat_articles = defaultdict(list)
    for row in data_rows:
        if not row or all(c is None for c in row):
            continue
        cat_name = str(row[1]).strip() if row[1] else None
        if not cat_name or cat_name == "Category":
            continue
        if cat_name not in cat_order:
            cat_order.append(cat_name)
        title   = str(row[2]).strip() if row[2] else ""
        full_url = str(row[4]).strip() if row[4] else ""
        keyword = str(row[5]).strip() if row[5] else title
        priority = clean_priority(row[8] if len(row) > 8 else None)
        affiliate = clean_affiliate(row[9] if len(row) > 9 else None)
        url_slug = extract_path_from_url(full_url)
        cat_articles[cat_name].append({
            "title": title,
            "url_slug": url_slug,
            "keyword": keyword,
            "affiliate": affiliate,
            "priority": priority,
        })

    # Build standard rows
    out = []
    for cat_name in cat_order:
        key = normalize_cat_key(cat_name)
        if key not in CATEGORY_MAP:
            print(f"  WARNING: unknown category '{cat_name}' — skipping")
            continue
        l1_path, _dir, id_prefix = CATEGORY_MAP[key]
        silo_header = f"{cat_name}\n{l1_path}"

        articles = cat_articles[cat_name]
        for idx, art in enumerate(articles, start=1):
            art_id = f"{id_prefix}{idx}"
            col0 = silo_header if idx == 1 else None
            out.append([col0, art_id, art["title"], art["url_slug"],
                        art["keyword"], art["affiliate"], art["priority"]])
    return out


def convert_format_c(rows):
    """Blue-mosque 6-col → standard 7-col rows."""
    # Row 0 IS the header in format C
    data_rows = rows[1:]

    cat_order = []
    cat_slugs = {}
    cat_articles = defaultdict(list)

    for row in data_rows:
        if not row or all(c is None for c in row):
            continue
        cat_name  = str(row[2]).strip() if len(row) > 2 and row[2] else None
        cat_slug  = str(row[3]).strip() if len(row) > 3 and row[3] else None
        art_slug  = str(row[4]).strip() if len(row) > 4 and row[4] else None
        title     = str(row[1]).strip() if len(row) > 1 and row[1] else ""

        if not cat_name or cat_name == "Category":
            continue
        if cat_name not in cat_order:
            cat_order.append(cat_name)
            cat_slugs[cat_name] = cat_slug

        url_slug = clean_url_slug(f"/{cat_slug}/{art_slug}/") if cat_slug and art_slug else "/"
        cat_articles[cat_name].append({
            "title": title,
            "url_slug": url_slug,
            "keyword": title,
            "affiliate": "—",
            "priority": "High",
        })

    out = []
    for cat_name in cat_order:
        key = normalize_cat_key(cat_name)
        if key in CATEGORY_MAP:
            l1_path = CATEGORY_MAP[key][0]
        else:
            # Fall back to building path from the slug
            cat_slug = cat_slugs.get(cat_name, cat_name.lower().replace(" ", "-"))
            l1_path = f"/{cat_slug}/"

        silo_header = f"{cat_name}\n{l1_path}"
        articles = cat_articles[cat_name]

        # Derive article ID prefix
        key = normalize_cat_key(cat_name)
        id_prefix = CATEGORY_MAP[key][2] if key in CATEGORY_MAP else cat_name[0].upper()

        for idx, art in enumerate(articles, start=1):
            art_id = f"{id_prefix}{idx}"
            col0 = silo_header if idx == 1 else None
            out.append([col0, art_id, art["title"], art["url_slug"],
                        art["keyword"], art["affiliate"], art["priority"]])
    return out


# ─── XLSX writer ─────────────────────────────────────────────────────────────

def write_standard_xlsx(path, data_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IA"

    header_font  = Font(bold=True, color=HEADER_FG)
    header_fill  = PatternFill("solid", fgColor=HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center")

    # Write header
    ws.append(HEADER)
    for cell in ws[1]:
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align

    # Write data
    for row in data_rows:
        ws.append(row)

    # Auto-width (approximate)
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                # Use longest line in cell (multiline silo headers)
                cell_len = max(len(line) for line in str(cell.value).split("\n"))
                max_len = max(max_len, cell_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    wb.save(path)


# ─── Main ─────────────────────────────────────────────────────────────────────

def process_file(xlsx_path):
    print(f"\nProcessing: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    all_rows = rows_from_ws(ws)

    # Find header row
    hi = None
    for i, row in enumerate(all_rows):
        vals = [str(c).strip() if c else "" for c in row]
        if "#" in vals:
            hi = i
            break
    if hi is None:
        print("  ERROR: could not find header row — skipping")
        return

    header_row = all_rows[hi]
    fmt = detect_format(header_row)
    print(f"  Detected format: {fmt}")

    if fmt == "A":
        print("  ✓ already standard — skipping")
        return

    if fmt == "B":
        data_rows = convert_format_b(all_rows)
    elif fmt == "C":
        data_rows = convert_format_c(all_rows)
    else:
        print(f"  ERROR: unsupported format '{fmt}' — skipping")
        return

    if not data_rows:
        print("  ERROR: conversion produced no rows — skipping")
        return

    # Backup original
    bak_path = xlsx_path + ".bak"
    if not os.path.exists(bak_path):
        shutil.copy2(xlsx_path, bak_path)
        print(f"  Backed up to: {bak_path}")
    else:
        print(f"  Backup already exists, overwriting file only")

    write_standard_xlsx(xlsx_path, data_rows)
    print(f"  ✓ Written {len(data_rows)} rows to: {xlsx_path}")


def main():
    if len(sys.argv) > 1:
        files = [os.path.abspath(sys.argv[1])]
    else:
        base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        pattern = os.path.join(base, "input", "*", "*.xlsx")
        files = [f for f in glob.glob(pattern) if not os.path.basename(f).startswith("~")]

    if not files:
        print("No XLSX files found.")
        sys.exit(0)

    print(f"Found {len(files)} XLSX file(s)")
    for f in sorted(files):
        process_file(f)

    print("\nDone.")


if __name__ == "__main__":
    main()
