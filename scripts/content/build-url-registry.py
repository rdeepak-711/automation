#!/usr/bin/env python3
"""
build-url-registry.py

Reads a standardised 7-col XLSX (output of standardize-xlsx.py) and builds
output/<site-slug>/url-registry.json.

Usage:
    python3 scripts/content/build-url-registry.py <site-slug>

Example:
    python3 scripts/content/build-url-registry.py hagia-sofia
"""

import glob
import json
import os
import sys
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

STANDARD_HEADER = ["#", "Category", "Article Title", "URL Slug", "Target Query", "Affiliate", "Priority"]

# Known mapping: category URL slug → actual directory name under input/<site>/
SLUG_TO_DIR_HINTS = {
    "tickets":        "tickets-tours",
    "tickets-tours":  "tickets-tours",
    "plan-your-visit": "plan-your-visit",
    "what-to-see":    "what-to-see",
}


def find_xlsx(site_dir):
    pattern = os.path.join(site_dir, "*.xlsx")
    matches = [f for f in glob.glob(pattern)
               if not os.path.basename(f).startswith("~")
               and not f.endswith(".bak")]
    if not matches:
        raise FileNotFoundError(f"No XLSX file found in: {site_dir}")
    return sorted(matches)[0]


def is_header_row(row):
    vals = [str(c).strip() if c else "" for c in row]
    return vals[:4] == ["#", "Category", "Article Title", "URL Slug"]


def is_silo_header(raw_col0):
    """Silo header rows have a newline in col 0: 'Display\n/path/'."""
    if raw_col0 is None:
        return False
    return "\n" in str(raw_col0)


def parse_silo_header(raw_col0):
    """Return (display_name, url_path) from 'Tickets & Tours\n/tickets/'."""
    parts = str(raw_col0).split("\n", 1)
    display = parts[0].strip()
    path    = parts[1].strip() if len(parts) > 1 else ""
    if path and not path.endswith("/"):
        path += "/"
    return display, path


def path_to_slug(path):
    """'/tickets/' → 'tickets'"""
    return path.strip("/").split("/")[0]


def ensure_slashes(path):
    if not path:
        return "/"
    path = str(path).strip().lstrip("/")  # strip any existing leading slashes
    path = "/" + path                      # add exactly one
    if not path.endswith("/"):
        path += "/"
    return path


def resolve_dir_name(cat_slug, site_dir):
    """Given a category slug from the URL, find the real directory under site_dir."""
    # 1. Check static hint map
    hint = SLUG_TO_DIR_HINTS.get(cat_slug)
    if hint and os.path.isdir(os.path.join(site_dir, hint)):
        return hint

    # 2. Exact match
    if os.path.isdir(os.path.join(site_dir, cat_slug)):
        return cat_slug

    # 3. Try all subdirs and see if any start with the slug
    try:
        subdirs = [d for d in os.listdir(site_dir)
                   if os.path.isdir(os.path.join(site_dir, d))]
    except OSError:
        return cat_slug

    for d in subdirs:
        if d.startswith(cat_slug):
            return d

    return cat_slug  # fallback — use slug directly


def load_xlsx(xlsx_path):
    """Parse standardised 7-col XLSX. Returns list of article dicts plus silo metadata."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Find header row
    header_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if is_header_row(row):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"Could not find standard header row in: {xlsx_path}\n"
                         "Run standardize-xlsx.py first.")

    silos   = {}   # cat_slug → {display, path, dir_name}
    articles = []  # list of dicts

    current_silo = None

    rows = list(ws.iter_rows(values_only=True))
    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        col0, cat_id, title, url_slug, keyword, affiliate, priority = (
            row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        )

        # New silo?
        if is_silo_header(col0):
            display, l1_path = parse_silo_header(col0)
            cat_slug = path_to_slug(l1_path)
            current_silo = {
                "display_name": display,
                "path": ensure_slashes(l1_path),
                "cat_slug": cat_slug,
            }
            silos[cat_slug] = current_silo

        # Article row?
        if cat_id and title:
            if current_silo is None:
                print(f"  WARNING: article '{title}' found before any silo header — skipping")
                continue
            url_slug = ensure_slashes(url_slug)
            slug_parts = url_slug.strip("/").split("/")
            art_slug   = slug_parts[-1] if slug_parts else ""
            cat_slug   = slug_parts[0] if slug_parts else current_silo["cat_slug"]

            articles.append({
                "category_id":   str(cat_id).strip(),
                "title":         str(title).strip(),
                "url_path":      url_slug,
                "slug":          art_slug,
                "category_slug": cat_slug,
                "keyword":       str(keyword).strip() if keyword else str(title).strip(),
                "affiliate":     str(affiliate).strip() if affiliate else "—",
                "priority":      str(priority).strip() if priority else "High",
            })

    return silos, articles


def load_json(path):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def extract_domain(homepage_cfg):
    return (homepage_cfg.get("domain")
            or homepage_cfg.get("site_url", "").replace("https://", "").replace("http://", "").rstrip("/")
            or "")


def build_expected_links(silos):
    """Build expected_links block — each L1 cross-links to all other L1s."""
    slugs = list(silos.keys())
    l1_cross = {}
    for s in slugs:
        l1_cross[s] = [o for o in slugs if o != s]
    l1_to_l0 = [silo["path"] for silo in silos.values()]
    return {"L1_cross": l1_cross, "L0_to_L1": l1_to_l0}


def build_registry(site_slug, site_dir, xlsx_path, l1_cfg, homepage_cfg):
    silos, articles = load_xlsx(xlsx_path)

    if not silos:
        raise ValueError("No silo headers found in XLSX. Cannot build registry.")

    # Resolve dir_names now that we know the site_dir
    for cat_slug, silo in silos.items():
        silo["dir_name"] = resolve_dir_name(cat_slug, site_dir)

    # ── pages ──────────────────────────────────────────────────────────────
    pages = {}

    # L0 homepage
    pages["/"] = {
        "type": "L0",
        "title": homepage_cfg.get("h1", "Homepage"),
        "wp_type": "page",
    }

    # L1 silo pages
    for cat_slug, silo in silos.items():
        pages[silo["path"]] = {
            "type":         "L1",
            "slug":         cat_slug,
            "dir_name":     silo["dir_name"],
            "display_name": silo["display_name"],
            "wp_type":      "page",
        }

    # L2 article pages
    for art in articles:
        cat_slug = art["category_slug"]
        dir_name = silos[cat_slug]["dir_name"] if cat_slug in silos else resolve_dir_name(cat_slug, site_dir)
        pages[art["url_path"]] = {
            "type":          "L2",
            "slug":          art["slug"],
            "title":         art["title"],
            "category_id":   art["category_id"],
            "category_slug": cat_slug,
            "dir_name":      dir_name,
            "wp_type":       "post",
            "affiliate":     art["affiliate"],
            "priority":      art["priority"],
        }

    # ── Scan actual HTML output files and register any missing slugs ──────────
    # Handles mismatches between XLSX slugs and MD-derived HTML filenames.
    output_l2_dir = os.path.join(
        os.path.dirname(os.path.dirname(site_dir)), "output", site_slug, "l2-articles"
    )
    if os.path.isdir(output_l2_dir):
        for dir_name in os.listdir(output_l2_dir):
            silo_dir = os.path.join(output_l2_dir, dir_name)
            if not os.path.isdir(silo_dir):
                continue
            # Use silo URL path from XLSX (e.g. tickets-tours dir → /tickets/ prefix)
            silo_url_prefix = None
            for cat_slug, silo in silos.items():
                if silo.get("dir_name") == dir_name:
                    silo_url_prefix = silo["path"].rstrip("/")
                    break
            if silo_url_prefix is None:
                silo_url_prefix = f"/{dir_name}"
            for html_file in glob.glob(os.path.join(silo_dir, "*.html")):
                slug = os.path.basename(html_file)[:-5]  # strip .html
                url_path = f"{silo_url_prefix}/{slug}/"
                if url_path not in pages:
                    pages[url_path] = {
                        "type":          "L2",
                        "slug":          slug,
                        "title":         slug.replace("-", " ").title(),
                        "category_slug": dir_name,
                        "dir_name":      dir_name,
                        "wp_type":       "post",
                        "affiliate":     "—",
                        "priority":      "Medium",
                        "_source":       "html_scan",
                    }

    # ── categories ─────────────────────────────────────────────────────────
    categories = {}
    for cat_slug, silo in silos.items():
        categories[cat_slug] = {
            "dir_name":     silo["dir_name"],
            "display_name": silo["display_name"],
        }

    # ── dir_to_url ──────────────────────────────────────────────────────────
    dir_to_url = {}
    for cat_slug, silo in silos.items():
        dir_to_url[silo["dir_name"]] = silo["path"].rstrip("/")

    # ── domain ─────────────────────────────────────────────────────────────
    domain = extract_domain(homepage_cfg) or f"{site_slug}.com"

    return {
        "domain":       domain,
        "site_slug":    site_slug,
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S"),
        "dir_to_url":   dir_to_url,
        "pages":        pages,
        "categories":   categories,
        "expected_links": build_expected_links(silos),
    }


def reconcile_ticket_l2_slugs(homepage_cfg_path, registry):
    """Patch tickets[].l2 in homepage-config.json to match registry slugs.

    populate-tickets.sh derives l2 slugs from tickets.md column 5, which can
    diverge from the XLSX-derived slugs in the registry. This runs after the
    registry is written and fixes any mismatches in-place.

    Matching strategy (in order):
      1. Exact title match (ticket title == article title)
      2. Current l2 slug is already a valid registry slug → keep it
      3. Slug-word overlap: registry slug whose words are a subset of the ticket slug words
      4. Warn unmatched tickets so they don't silently stay wrong
    """
    cfg = load_json(homepage_cfg_path)
    tickets = cfg.get("tickets", [])
    if not tickets:
        return

    # Build lookups from registry L2 pages
    title_to_slug = {}
    valid_slugs = set()
    for page in registry["pages"].values():
        if page.get("type") == "L2" and page.get("slug"):
            valid_slugs.add(page["slug"])
            if page.get("title"):
                title_to_slug[page["title"].lower()] = page["slug"]

    def _slug_words(s):
        return set(s.replace("-", " ").split())

    fixes = []
    unmatched = []
    for ticket in tickets:
        title = ticket.get("title", "")
        current_l2 = ticket.get("l2", "")

        # 1. Exact title match
        registry_slug = title_to_slug.get(title.lower())

        # 2. Current slug already valid
        if not registry_slug and current_l2 in valid_slugs:
            continue

        # 3. Slug-word overlap: find registry slug whose words ⊆ ticket slug words
        if not registry_slug:
            ticket_words = _slug_words(current_l2)
            candidates = [s for s in valid_slugs if _slug_words(s) <= ticket_words]
            if len(candidates) == 1:
                registry_slug = candidates[0]

        if registry_slug and current_l2 != registry_slug:
            fixes.append(f"    {title[:50]}: {current_l2!r} → {registry_slug!r}")
            ticket["l2"] = registry_slug
        elif not registry_slug and current_l2 not in valid_slugs:
            unmatched.append(f"    {title[:50]}: {current_l2!r} not in registry")

    if fixes:
        with open(homepage_cfg_path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2, ensure_ascii=False)
        print(f"  ✓ Reconciled {len(fixes)} ticket l2 slug(s) with registry:")
        for fix in fixes:
            print(fix)
    else:
        print("  ✓ Ticket l2 slugs already match registry — no changes needed")

    if unmatched:
        print(f"  ⚠ {len(unmatched)} ticket(s) could not be matched to registry slugs:")
        for u in unmatched:
            print(u)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    site_slug = sys.argv[1].strip("/")
    base_dir  = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    site_dir  = os.path.join(base_dir, "input", site_slug)

    if not os.path.isdir(site_dir):
        print(f"ERROR: Site directory not found: {site_dir}")
        sys.exit(1)

    # Locate XLSX
    try:
        xlsx_path = find_xlsx(site_dir)
    except FileNotFoundError as e:
        print(f"ERROR: {e}")
        sys.exit(1)
    print(f"Using XLSX: {xlsx_path}")

    # Load optional config files
    homepage_cfg_path = os.path.join(site_dir, "homepage-config.json")
    l1_cfg       = load_json(os.path.join(site_dir, "l1-config.json"))
    homepage_cfg = load_json(homepage_cfg_path)

    # Build registry
    try:
        registry = build_registry(site_slug, site_dir, xlsx_path, l1_cfg, homepage_cfg)
    except ValueError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    # Write output
    out_dir  = os.path.join(base_dir, "output", site_slug)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "url-registry.json")

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(registry, f, indent=2, ensure_ascii=False)

    l2_count = sum(1 for p in registry["pages"].values() if p["type"] == "L2")
    l1_count = sum(1 for p in registry["pages"].values() if p["type"] == "L1")
    print(f"✓ Written url-registry.json")
    print(f"  Domain   : {registry['domain']}")
    print(f"  L1 silos : {l1_count}")
    print(f"  L2 pages : {l2_count}")
    print(f"  Output   : {out_path}")

    # Reconcile tickets[].l2 slugs with registry (fixes mismatches from Step 18)
    if os.path.exists(homepage_cfg_path) and homepage_cfg.get("tickets"):
        reconcile_ticket_l2_slugs(homepage_cfg_path, registry)


if __name__ == "__main__":
    main()
