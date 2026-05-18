#!/usr/bin/env python3
"""
Fix card images on the homepage and L1 pages — replaces placeholder/wrong
images with each post's actual featured image, matched by the nearest
internal article href in the card.

Pass 0 (XLSX slug audit): if xlsx_path provided, checks every internal card
href against XLSX canonical slugs and fixes mismatches before touching images.

Handles ALL image classes:
  att-hero__img       — page hero banner
  att-article-card__img — 3-col article grid cards
  att-featured__img   — top highlights / featured horizontal cards
  att-ticket__img     — ticket grid cards (homepage + L1)
  att-highlight__img  — highlight grid (homepage)
  att-crosslink__img  — cross-link cards between silos

Strategy: for each placeholder img, scan forward up to 2000 chars to find
the nearest internal /category/slug/ href, then replace with that post's
featured image. This works for all card types regardless of whether the
img is inside or outside the <a> tag.

Hero images and crosslink images use the per-page hero URLs passed as args 3-6.
Crosslinks are always updated to the designated hero regardless of current src.

Usage:
    python3 fix-l1-card-images.py <wp_path> <site_url> [homepage_img] [tickets_img] [plan_your_visit_img] [what_to_see_img] [xlsx_path]

Example:
    python3 fix-l1-card-images.py /home1/kzrmeomy/public_html/website_ce6ca565 https://auschwitz-guide.com
    python3 fix-l1-card-images.py /home1/dpskbcmy/public_html/website_windsor https://windsorcastle-guide.com https://... https://... https://... https://... /tmp/site.xlsx
"""

import subprocess, re, sys, os, tempfile

PLACEHOLDER_GIF = "data:image/gif;base64,R0lGODlhAQABAAAAACH5BAEKAAEALAAAAAABAAEAAAICTAEAOw=="

def usage():
    print(__doc__)
    sys.exit(1)

if len(sys.argv) < 3:
    usage()

WP_PATH  = sys.argv[1].rstrip("/")
SITE_URL = sys.argv[2].rstrip("/")
SITE_DOMAIN = re.sub(r"^https?://(?:www\.)?", "", SITE_URL)

# Optional per-page hero images (args 3-6)
PAGE_HERO_IMGS = {}
if len(sys.argv) >= 7:
    for slug, url in zip(
        ["homepage", "tickets", "plan-your-visit", "what-to-see"],
        sys.argv[3:7]
    ):
        if url.strip():
            PAGE_HERO_IMGS[slug] = url.strip()

if PAGE_HERO_IMGS:
    print(f"Hero images provided for: {list(PAGE_HERO_IMGS.keys())}")
else:
    print("No hero images provided — will use post featured image fallback for heroes/crosslinks")

# ── XLSX slug audit (optional arg 7) ─────────────────────────────────────────
XLSX_PATH = sys.argv[7].strip() if len(sys.argv) >= 8 else None

# xlsx_valid_slugs: category → set of canonical slugs from XLSX
xlsx_valid_slugs = {}   # e.g. {"tickets": {"park-guell-admission-ticket", ...}, ...}

if XLSX_PATH and os.path.exists(XLSX_PATH):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        cat_col  = next((i for i, h in enumerate(headers) if h.lower() == "category"), None)
        slug_col = next((i for i, h in enumerate(headers) if "slug" in h.lower()), None)
        if cat_col is not None and slug_col is not None:
            CAT_PREFIX = {"T": "tickets", "P": "plan-your-visit", "W": "what-to-see"}
            for row in ws.iter_rows(min_row=2, values_only=True):
                cat_raw  = str(row[cat_col]).strip()  if row[cat_col]  else ""
                slug_raw = str(row[slug_col]).strip() if row[slug_col] else ""
                if not cat_raw or not slug_raw:
                    continue
                prefix = cat_raw[0].upper()
                silo = CAT_PREFIX.get(prefix)
                if not silo:
                    continue
                # Extract just the slug portion: /tickets/my-slug/ → my-slug
                parts = [p for p in slug_raw.strip("/").split("/") if p]
                if len(parts) >= 2:
                    xlsx_valid_slugs.setdefault(silo, set()).add(parts[-1])
            total_xlsx = sum(len(v) for v in xlsx_valid_slugs.values())
            print(f"XLSX loaded: {total_xlsx} slugs across {list(xlsx_valid_slugs.keys())}")
        else:
            print("WARNING: XLSX missing Category or Slug column — skipping slug audit")
    except ImportError:
        print("WARNING: openpyxl not installed — skipping slug audit")
    except Exception as e:
        print(f"WARNING: XLSX load failed ({e}) — skipping slug audit")
elif XLSX_PATH:
    print(f"WARNING: XLSX not found at {XLSX_PATH} — skipping slug audit")

def _slug_words(text):
    stop = {"the", "a", "an", "and", "or", "of", "to", "in", "for", "with", "from", "by", "at", "on", "is", "amp"}
    return set(w for w in re.sub(r"[^a-z0-9\s]", "", text.replace("-", " ").lower()).split()
               if w not in stop and len(w) > 2)

def find_best_xlsx_slug(wrong_slug, silo):
    """Word-overlap match wrong_slug against XLSX candidates for silo. Returns best match or None."""
    candidates = xlsx_valid_slugs.get(silo, set())
    if not candidates:
        return None
    ww = _slug_words(wrong_slug)
    best, best_score = None, 0
    for cand in candidates:
        score = len(ww & _slug_words(cand))
        if score > best_score:
            best_score, best = score, cand
    return best if best_score >= 2 else None

def fix_card_slugs(content, page_label):
    """
    Pass 0: find every internal /silo/slug/ href in content.
    If slug not in XLSX valid set for that silo, find best match and replace.
    Returns (new_content, fixes_made).
    """
    if not xlsx_valid_slugs:
        return content, 0

    fixes = 0
    href_pat = re.compile(r'href="(/([a-z0-9-]+)/([a-z0-9-]+)/)"')

    def replacer(m):
        nonlocal fixes
        full_path, silo, slug = m.group(1), m.group(2), m.group(3)
        valid = xlsx_valid_slugs.get(silo)
        if valid is None:
            return m.group(0)  # silo not in XLSX (e.g. homepage anchor) — skip
        if slug in valid:
            return m.group(0)  # correct — leave alone
        best = find_best_xlsx_slug(slug, silo)
        if best:
            fixes += 1
            new_path = f"/{silo}/{best}/"
            print(f"  ✎ slug fix [{silo}]: {slug} → {best}")
            return f'href="{new_path}"'
        else:
            print(f"  ⚠ slug not in XLSX, no match found [{silo}]: {slug}")
            return m.group(0)

    new_content = href_pat.sub(replacer, content)
    return new_content, fixes

# Matches internal hrefs with 2+ path segments: /silo/slug/ or https://domain/silo/slug/
# Trailing slash is optional (some cards omit it)
ARTICLE_HREF = re.compile(
    r'href="(?:https?://(?:www\.)?' + re.escape(SITE_DOMAIN) + r')?(/[a-z0-9-]+/[a-z0-9-]+/?)"'
)

def wp(*args):
    result = subprocess.run(
        ["wp", "--path=" + WP_PATH] + list(args),
        capture_output=True, text=True
    )
    return result.stdout.strip()

# ── Build slug → article header image URL map ────────────────────────────────
# Source of truth: the att-article-header__image src inside each post's content.
# This is the image the article itself declares, not the WP featured image meta.
print("Building slug → image map from article header images...")
posts_csv = wp("post", "list", "--post_type=post", "--format=csv",
               "--fields=ID,post_name", "--posts_per_page=500")
slug_img = {}          # post_slug → image URL
cat_first_img = {}     # category_slug → first available image URL in that cat

for line in posts_csv.splitlines()[1:]:
    parts = line.split(",", 1)
    if len(parts) != 2:
        continue
    post_id, post_slug = parts[0].strip(), parts[1].strip()
    content = wp("post", "get", post_id, "--field=post_content")
    m = re.search(r'class="att-article-header__image"\s+src="([^"]+)"', content)
    if not m:
        # Also try src before class attribute order
        m = re.search(r'src="([^"]+)"\s+[^>]*class="att-article-header__image"', content)
    if not m:
        continue
    img_url = m.group(1)
    if not img_url or img_url.startswith("data:"):
        continue
    slug_img[post_slug] = img_url
    # Track first available image per category (fallback when no hero provided)
    cats = wp("post", "term", "list", post_id, "category",
              "--fields=slug", "--format=csv").splitlines()
    for cat in cats[1:]:  # skip header
        cat = cat.strip()
        if cat and cat not in cat_first_img:
            cat_first_img[cat] = img_url

print(f"  {len(slug_img)} posts with article header images")
print(f"  Categories with images: {list(cat_first_img.keys())}")
if not slug_img:
    print("  WARNING: No posts have featured images — nothing to do")
    sys.exit(0)

# First available image as last-resort fallback
_first_img = next(iter(slug_img.values()))

def get_slug_from_href(href):
    """Extract last path segment from an internal href."""
    parts = [p for p in href.strip("/").split("/") if p]
    return parts[-1] if len(parts) >= 2 else None

def get_cat_from_href(href):
    """Extract first path segment (category) from an internal href."""
    parts = [p for p in href.strip("/").split("/") if p]
    return parts[0] if parts else None

# L1-page slugs — crosslinks pointing to these use hero or silo fallback image
L1_SLUGS = {"tickets", "tickets-tours", "plan-your-visit", "what-to-see",
            "homepage", "home", "tickets-and-tours"}

def find_best_image(text_window, allow_l1_fallback=True):
    """
    Scan text_window for the nearest internal href, try to match to a post image.
    - /category/slug/  → direct post lookup
    - /category/       → hero image if provided, else first image in that category silo
      (only when allow_l1_fallback=True — crosslinks only; not for article/highlight cards)
    Returns (label, img_url) or (None, None).
    """
    # First pass: try article links (2 segments) → direct post match
    for m in ARTICLE_HREF.finditer(text_window):
        slug = get_slug_from_href(m.group(1))
        if slug and slug in slug_img:
            return slug, slug_img[slug]

    if not allow_l1_fallback:
        return None, None

    # Second pass: try any internal href — including L1 page links (1 segment)
    any_href = re.compile(r'href="(?:https?://(?:www\.)?' + re.escape(SITE_DOMAIN) + r')?(/[a-z0-9-]+/?)"')
    for m in any_href.finditer(text_window):
        cat = get_cat_from_href(m.group(1))
        if not cat:
            continue
        if cat not in L1_SLUGS:
            continue  # skip non-L1 single-segment hrefs — not reliable
        # Use designated hero image if provided, else silo fallback
        if cat in PAGE_HERO_IMGS:
            return f"{cat} (hero)", PAGE_HERO_IMGS[cat]
        if cat in cat_first_img:
            return f"{cat} (silo fallback)", cat_first_img[cat]

    return None, None

def fix_placeholders_in_content(content, page_label, page_slug):
    """
    Replace all placeholder images in content.
    Scans forward 2000 chars after each placeholder to find its article href.
    Returns (new_content, fixed_count, remaining_count).
    """
    # Hero image for this specific page
    hero_img = PAGE_HERO_IMGS.get(page_slug, _first_img)

    fixed = 0
    offset = 0
    new_content = content

    while True:
        pos = new_content.find(PLACEHOLDER_GIF, offset)
        if pos == -1:
            break

        # Get image class for logging — look back up to 600 chars (multi-line img tags)
        before = new_content[max(0, pos-600):pos]
        cls_match = re.search(r'class="([^"]*att-[a-z_-]+__img[^"]*)"', before)
        cls = cls_match.group(1) if cls_match else "unknown"

        # Scan forward to find article href
        # L1 hero fallback only for crosslink images — not article/highlight cards
        _article_classes = ("att-article-card__img", "att-highlight__img", "att-featured__img", "att-ticket__img")
        _allow_l1 = not any(c in cls for c in _article_classes)
        window = new_content[pos:pos + 2000]
        slug, img_url = find_best_image(window, allow_l1_fallback=_allow_l1)

        if img_url:
            new_content = new_content[:pos] + img_url + new_content[pos + len(PLACEHOLDER_GIF):]
            fixed += 1
            print(f"  ✓ {cls}: {slug} → {img_url.split('/')[-1]}")
            # Don't advance offset — string is shorter now, recheck same pos
        else:
            # No article link found — use hero fallback
            if "att-hero__img" in cls:
                new_content = new_content[:pos] + hero_img + new_content[pos + len(PLACEHOLDER_GIF):]
                fixed += 1
                print(f"  ✓ {cls}: hero → {hero_img.split('/')[-1]}")
            else:
                print(f"  ✗ {cls}: no matching post with featured image")
                offset = pos + len(PLACEHOLDER_GIF)

    remaining = new_content.count(PLACEHOLDER_GIF)
    return new_content, fixed, remaining

def fix_crosslink_images(content):
    """
    Update att-crosslink__img srcs to use designated hero images.
    Runs on every page regardless of placeholder presence — ensures
    crosslinks always show the correct silo hero image.
    Only updates if PAGE_HERO_IMGS is populated.
    """
    if not PAGE_HERO_IMGS:
        return content, 0

    fixed = 0

    def replacer(m):
        nonlocal fixed
        block = m.group(0)
        href_m = re.search(r'href="/([\w-]+)/?"', block)
        if not href_m:
            return block
        silo = href_m.group(1)
        if silo not in PAGE_HERO_IMGS:
            return block
        new_block = re.sub(
            r'(<img[^>]*class="att-crosslink__img"[^>]*)src="[^"]*"',
            lambda im: im.group(1) + f'src="{PAGE_HERO_IMGS[silo]}"',
            block
        )
        if new_block != block:
            fixed += 1
            print(f"  ✓ att-crosslink__img → {silo}: {PAGE_HERO_IMGS[silo].split('/')[-1]}")
        return new_block

    new_content = re.sub(
        r'<div class="att-crosslink">.*?</div>\s*</div>',
        replacer, content, flags=re.DOTALL
    )
    return new_content, fixed

def fix_hero_image(content, page_slug):
    """
    Update att-hero__img src to the designated hero image for this page.
    Runs regardless of placeholder presence — ensures hero is always correct.
    Only updates if PAGE_HERO_IMGS has an entry for this page.
    """
    if page_slug not in PAGE_HERO_IMGS:
        return content, 0

    hero_url = PAGE_HERO_IMGS[page_slug]
    new_content = re.sub(
        r'(<img[^>]*class="att-hero__img"[^>]*)src="[^"]*"',
        lambda m: m.group(1) + f'src="{hero_url}"',
        content
    )
    # Also handle src-before-class order
    new_content = re.sub(
        r'(<img[^>]*)src="[^"]*"([^>]*class="att-hero__img")',
        lambda m: m.group(1) + f'src="{hero_url}"' + m.group(2),
        new_content
    )
    changed = 1 if new_content != content else 0
    if changed:
        print(f"  ✓ att-hero__img → {hero_url.split('/')[-1]}")
    return new_content, changed

def set_featured_image(page_id, page_slug):
    """Set the WP featured image (_thumbnail_id) for an L1 page from PAGE_HERO_IMGS."""
    hero_url = PAGE_HERO_IMGS.get(page_slug)
    if not hero_url:
        return
    filename = hero_url.split("/")[-1]
    att_id = wp("attachment", "url-to-id", hero_url)
    if not att_id or not att_id.isdigit():
        # Image not yet in media library — import from local uploads path
        # Derive file path from URL: strip scheme+domain, prepend WP_PATH
        path_part = re.sub(r"^https?://[^/]+", "", hero_url)
        local_path = WP_PATH + path_part
        print(f"  Importing {filename} into media library...")
        att_id = wp("media", "import", local_path, "--porcelain")
    if not att_id or not att_id.isdigit():
        print(f"  ✗ featured image: could not register {filename}")
        return
    wp("post", "meta", "update", page_id, "_thumbnail_id", att_id)
    print(f"  ✓ featured image set → {filename} (ID {att_id})")

def save_page(page_id, content):
    tmp = f"/tmp/wp_page_{page_id}.html"
    with open(tmp, "w") as f:
        f.write(content)
    proc = subprocess.run(
        ["wp", "--path=" + WP_PATH, "eval",
         f"wp_update_post(['ID'=>{page_id},'post_content'=>file_get_contents('{tmp}')]);"],
        capture_output=True, text=True
    )
    try:
        os.unlink(tmp)
    except OSError:
        pass
    if proc.returncode == 0:
        print(f"  Saved OK")
    else:
        print(f"  SAVE FAILED: {proc.stderr[:300]}")

# ── Get all pages to process ─────────────────────────────────────────────────
pages_csv = wp("post", "list", "--post_type=page", "--format=csv",
               "--fields=ID,post_name", "--posts_per_page=50")

SKIP_SLUGS = {"about-us", "contact-us", "privacy-policy", "terms", "sitemap"}
pages_to_fix = []
for line in pages_csv.splitlines()[1:]:
    parts = line.split(",", 1)
    if len(parts) != 2:
        continue
    page_id, page_slug = parts[0].strip(), parts[1].strip()
    if page_slug not in SKIP_SLUGS:
        pages_to_fix.append((page_id, page_slug))

print(f"\nPages to process: {[s for _, s in pages_to_fix]}")

total_fixed = 0
total_remaining = 0

for page_id, page_slug in pages_to_fix:
    print(f"\n=== Page {page_id} ({page_slug}) ===")
    content = wp("post", "get", page_id, "--field=post_content")
    changed = False

    # Pass 0: fix card slugs against XLSX canonical slugs
    content, slug_fixes = fix_card_slugs(content, page_slug)
    if slug_fixes:
        print(f"  Slug fixes: {slug_fixes}")
        changed = True

    # Pass 1: fix placeholder GIFs
    placeholder_count = content.count(PLACEHOLDER_GIF)
    if placeholder_count > 0:
        print(f"  Placeholders found: {placeholder_count}")
        content, fixed, remaining = fix_placeholders_in_content(content, page_slug, page_slug)
        total_fixed += fixed
        total_remaining += remaining
        print(f"  Fixed: {fixed} | Remaining: {remaining}")
        if fixed > 0:
            changed = True

    # Pass 2: force-set hero image to designated URL
    content, hero_changed = fix_hero_image(content, page_slug)
    if hero_changed:
        changed = True

    # Pass 3: force-set crosslink images to silo hero URLs
    content, cross_fixed = fix_crosslink_images(content)
    if cross_fixed:
        total_fixed += cross_fixed
        changed = True

    if changed:
        save_page(page_id, content)
    else:
        print(f"  Nothing changed — skipping save")

    # Pass 4: set WP featured image for this page
    set_featured_image(page_id, page_slug)

# ── Flush all caches ─────────────────────────────────────────────────────────
print("\nFlushing caches...")
wp("cache", "flush")
wp("transient", "delete", "--all")
subprocess.run(
    ["bash", "-c",
     f"wp --path={WP_PATH} eval 'if(function_exists(\"rocket_clean_domain\")){{rocket_clean_domain();}} "
     f"if(function_exists(\"rocket_clean_minify\")){{rocket_clean_minify();}} opcache_reset();' 2>/dev/null || true"],
    capture_output=True
)
subprocess.run(
    ["bash", "-c",
     f"find {WP_PATH}/wp-content/cache/ -name '*.html' -delete 2>/dev/null; "
     f"find {WP_PATH}/wp-content/cache/ -name '*.html.gz' -delete 2>/dev/null"],
    capture_output=True
)

print(f"\n{'='*50}")
print(f"Total fixed: {total_fixed}")
print(f"Total remaining (no featured image): {total_remaining}")
print("Done.")
