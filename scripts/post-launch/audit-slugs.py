#!/usr/bin/env python3
"""
Audit and fix WordPress post slugs against the site's XLSX article index.

Compares the 'Article Slug' column in the XLSX against live WP post_names.
Mismatched slugs are renamed in WP and all internal URLs are search-replaced
across the entire database.

Usage (run via audit-slugs.sh — do not call directly):
    python3 audit-slugs.py <wp_path> <xlsx_path> [--fix]

Modes:
    --audit   (default) report only, no changes
    --fix     rename mismatched slugs in WP and search-replace all URLs
"""

import subprocess, re, sys, os

def usage():
    print(__doc__)
    sys.exit(1)

if len(sys.argv) < 3:
    usage()

WP_PATH   = sys.argv[1].rstrip("/")
XLSX_PATH = sys.argv[2]
FIX_MODE  = "--fix" in sys.argv

def wp(*args):
    r = subprocess.run(["wp", "--path=" + WP_PATH] + list(args), capture_output=True, text=True)
    return r.stdout.strip()

# ── Load XLSX slugs ───────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

# Detect header row — find 'Article Slug' column
headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
SLUG_COL_NAMES = ["Article Slug", "URL Slug", "Slug"]
try:
    cat_col = headers.index("Category")
    num_col = headers.index("#")
    slug_col = next((headers.index(n) for n in SLUG_COL_NAMES if n in headers), None)
    if slug_col is None:
        raise ValueError(f"No slug column found (tried: {SLUG_COL_NAMES})")
except ValueError as e:
    print(f"ERROR: Could not find required column: {e}")
    print(f"  Found columns: {headers}")
    sys.exit(1)

print(f"  Slug column: '{headers[slug_col]}'")

xlsx_entries = []  # list of (num, category, xlsx_slug)
current_category = ""
for row in ws.iter_rows(min_row=2, values_only=True):
    raw_num  = row[num_col]
    category = row[cat_col]
    raw_slug = row[slug_col]
    if not raw_slug:
        continue

    # Detect category header rows: num cell contains "Category\n/prefix/" or similar
    if raw_num and "\n" in str(raw_num):
        current_category = str(raw_num).split("\n")[0].strip()

    # Use category column value as num if # column is None or a header string
    if raw_num is None or (isinstance(raw_num, str) and "\n" in raw_num):
        num = str(category).strip() if category else "?"
    else:
        num = raw_num

    # Resolve category label
    cat_label = current_category if current_category else str(category).strip()

    # Extract just the final path segment from slugs like /tickets/entry-ticket/
    slug = str(raw_slug).strip().strip("/")
    if "/" in slug:
        slug = slug.split("/")[-1]

    if slug:
        xlsx_entries.append((num, cat_label, slug))

print(f"XLSX: {len(xlsx_entries)} article slugs loaded from {os.path.basename(XLSX_PATH)}")

# ── Load WP post slugs ────────────────────────────────────────────────────────
posts_csv = wp("post", "list", "--post_type=post", "--post_status=publish",
               "--format=csv", "--fields=ID,post_name", "--posts_per_page=500")

wp_slug_to_id = {}
for line in posts_csv.splitlines()[1:]:
    parts = line.split(",", 1)
    if len(parts) == 2:
        wp_slug_to_id[parts[1].strip()] = parts[0].strip()

print(f"WP:   {len(wp_slug_to_id)} published posts found")

# ── Compare ───────────────────────────────────────────────────────────────────
matched    = []
mismatched = []  # (num, cat, xlsx_slug, wp_slug, post_id)
missing    = []  # (num, cat, xlsx_slug) — no close match in WP at all

wp_slugs = set(wp_slug_to_id.keys())

for num, cat, xlsx_slug in xlsx_entries:
    if xlsx_slug in wp_slugs:
        matched.append(xlsx_slug)
        continue

    # Find closest WP slug: WP slug contains xlsx_slug or starts with it
    close = [s for s in wp_slugs if xlsx_slug in s or s.startswith(xlsx_slug)]
    if close:
        # Pick shortest match
        wp_slug = sorted(close, key=len)[0]
        mismatched.append((num, cat, xlsx_slug, wp_slug, wp_slug_to_id[wp_slug]))
    else:
        missing.append((num, cat, xlsx_slug))

# ── Report ────────────────────────────────────────────────────────────────────
print()
print(f"{'#':<4} {'Category':<20} {'XLSX Slug':<45} {'Status'}")
print("-" * 100)

for xlsx_slug in matched:
    entry = next(e for e in xlsx_entries if e[2] == xlsx_slug)
    print(f"{entry[0]:<4} {entry[1]:<20} {xlsx_slug:<45} ✅ MATCH")

for num, cat, xlsx_slug, wp_slug, post_id in mismatched:
    print(f"{num:<4} {cat:<20} {xlsx_slug:<45} ⚠️  MISMATCH → WP: {wp_slug}")

for num, cat, xlsx_slug in missing:
    print(f"{num:<4} {cat:<20} {xlsx_slug:<45} ❌ MISSING in WP")

print()
print(f"✅ Matched:    {len(matched)}")
print(f"⚠️  Mismatched: {len(mismatched)}")
print(f"❌ Missing:    {len(missing)}")

extra_wp = wp_slugs - {e[2] for e in xlsx_entries}
if extra_wp:
    print(f"\nWP posts not in XLSX ({len(extra_wp)}):")
    for s in sorted(extra_wp):
        print(f"  → {s}")

if not mismatched:
    print("\nNothing to fix.")
    sys.exit(0)

if not FIX_MODE:
    print(f"\nRun with --fix to rename the {len(mismatched)} mismatched slugs in WP.")
    sys.exit(0)

# ── Fix: rename WP slugs + search-replace URLs ────────────────────────────────
print(f"\n{'='*60}")
print("FIXING: renaming WP slugs and replacing all internal URLs")
print('='*60)

for num, cat, xlsx_slug, wp_slug, post_id in mismatched:
    wp("post", "update", post_id, f"--post_name={xlsx_slug}", "--quiet")
    print(f"  ✅ ID {post_id}: {wp_slug} → {xlsx_slug}")

print()
print("Search-replacing old URLs across all tables...")
for num, cat, xlsx_slug, wp_slug, post_id in mismatched:
    result = subprocess.run(
        ["wp", "--path=" + WP_PATH, "search-replace",
         wp_slug, xlsx_slug, "--all-tables", "--report-changed-only"],
        capture_output=True, text=True
    )
    changed_lines = [l for l in result.stdout.splitlines()
                     if l.strip() and not l.startswith("Success")]
    count = sum(
        int(m.group(1)) for l in changed_lines
        if (m := re.search(r'\b(\d+)\b', l))
    )
    print(f"  ✅ {wp_slug} → {xlsx_slug}  ({count} replacements)")

print()
wp("cache", "flush")
wp("transient", "delete", "--all")
print("Cache flushed.")
print(f"\nDone. {len(mismatched)} slugs renamed and URLs updated.")
