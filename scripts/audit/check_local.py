#!/usr/bin/env python3
"""Layer 2b: CHECK_LOCAL — local file vs snapshot diff.

Covers:
  - XLSX manifest vs live slugs
  - tickets.md integrity
  - input/<slug>/images/favicon + logo presence + live site_icon
  - md → html parity in output/<slug>/
Output: violations list (merged into main violations.json by audit.sh)

Usage: python3 check_local.py <snapshot.json> [--out local-violations.json]
"""
import sys, json, re, argparse, os
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


HOSTNAME_TO_INPUT_SLUG = {
    'topkapipalace-guide.com':      'topkapi-palace',
    'bluemosque-guide.com':         'blue-mosque',
    'hagiasophia-guide.com':        'hagia-sofia',
    'montsaintmichel-guide.com':    'mont-saint-michel',
    'vangoghmuseum-guide.com':      'van-gogh',
    'plitvicelakes-guide.com':      'plitvice-lakes',
    'angkorwat-guide.com':          'angkor-wat',
    'pyramidsofgiza-guide.com':     'pyramids-of-giza',
    'amsterdamcanalcruise-guide.com': 'amsterdam-canal-cruise',
    'guide-stonehenge.com':         'stonehenge',
    'museodelprado-guide.com':      'museo-del-prado',
}


def v(severity, vid, category, target, message, **extra):
    return {
        'id':       vid,
        'severity': severity,
        'category': category,
        'target':   target,
        'message':  message,
        **extra,
    }


def check_xlsx_manifest(input_dir, snap):
    """Diff XLSX sheet IA col 4 slugs vs live post slugs."""
    out = []
    if not HAS_OPENPYXL:
        out.append(v('low', 'openpyxl-missing', 'local-manifest',
                     {'type': 'env'}, 'openpyxl not installed — skipping XLSX diff'))
        return out
    xlsx_files = list(input_dir.glob('*.xlsx'))
    xlsx_files = [f for f in xlsx_files if not f.name.endswith('.bak')]
    if not xlsx_files:
        out.append(v('medium', 'xlsx-manifest-missing', 'local-manifest',
                     {'type': 'file', 'dir': str(input_dir)},
                     f"no *.xlsx manifest in {input_dir}"))
        return out

    xlsx_slugs = set()
    xlsx_path = xlsx_files[0]
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb['IA'] if 'IA' in wb.sheetnames else wb.active

        # Detect header row to find col indices for URL Slug (D) and Live URL (H)
        headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        slug_col = next((i for i, h in enumerate(headers) if 'url slug' in h or h == 'url slug'), 3)
        live_col = next((i for i, h in enumerate(headers) if 'live url' in h), None)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            # Prefer Live URL col if present, fall back to URL Slug col
            slug = row[live_col] if (live_col is not None and live_col < len(row) and row[live_col]) else (row[slug_col] if slug_col < len(row) else None)
            if not slug or not isinstance(slug, str):
                continue
            parts = [p for p in slug.strip('/').split('/') if p]
            if len(parts) >= 2:
                xlsx_slugs.add(parts[-1])
    except Exception as e:
        out.append(v('medium', 'xlsx-read-error', 'local-manifest',
                     {'type': 'file', 'path': str(xlsx_path)},
                     f"failed to read XLSX: {e}"))
        return out

    live_slugs = {p['slug'] for p in snap['posts']} | {p['slug'] for p in snap['pages']}

    missing = xlsx_slugs - live_slugs  # in manifest, not live → unpublished
    extra   = {s for s in live_slugs if s not in xlsx_slugs} - {p['slug'] for p in snap['pages']}  # orphan posts only

    for slug in sorted(missing):
        out.append(v('high', 'manifest-unpublished', 'local-manifest',
                     {'type': 'slug', 'slug': slug},
                     f"slug '{slug}' in XLSX manifest but NOT live on site",
                     source_xlsx=str(xlsx_path), fixable=False))
    for slug in sorted(extra):
        out.append(v('medium', 'manifest-orphan', 'local-manifest',
                     {'type': 'slug', 'slug': slug},
                     f"slug '{slug}' live on site but NOT in XLSX manifest (orphan)",
                     source_xlsx=str(xlsx_path), fixable=False))
    return out


def check_tickets_md(input_dir, snap):
    """Validate tickets.md structure + URL uniqueness."""
    out = []
    tm = input_dir / 'tickets.md'
    if not tm.exists():
        out.append(v('low', 'tickets-md-missing', 'local-tickets',
                     {'type': 'file', 'path': str(tm)},
                     "tickets.md not present (skip if site has no paid products)"))
        return out

    seen_ids  = {}
    seen_urls = {}
    bad_rows  = []
    for lineno, raw in enumerate(open(tm), start=1):
        line = raw.strip()
        if not line or line.startswith('#'):
            continue
        parts = [p.strip() for p in line.split('|')]
        if len(parts) < 3:
            bad_rows.append((lineno, "fewer than 3 pipe-delimited fields"))
            continue
        tid, name, url = parts[0], parts[1], parts[2]
        platform = parts[3] if len(parts) > 3 else ''
        if not tid or not re.match(r'^T\d+$', tid):
            bad_rows.append((lineno, f"bad id {tid!r} — expected T<n>"))
        if not name:
            bad_rows.append((lineno, "empty name"))
        if not re.match(r'^https?://', url):
            bad_rows.append((lineno, f"bad url {url!r}"))
        if tid in seen_ids:
            bad_rows.append((lineno, f"duplicate id {tid} (also on line {seen_ids[tid]})"))
        else:
            seen_ids[tid] = lineno
        if url in seen_urls:
            bad_rows.append((lineno, f"duplicate url (also on line {seen_urls[url]})"))
        else:
            seen_urls[url] = lineno

    for lineno, msg in bad_rows:
        out.append(v('medium', 'tickets-md-bad-row', 'local-tickets',
                     {'type': 'file_line', 'path': str(tm), 'line': lineno},
                     f"tickets.md line {lineno}: {msg}"))
    return out


def check_input_assets(input_dir, snap):
    """Favicon + logo file presence locally; site_icon presence on live site."""
    out = []
    images_dir = input_dir / 'images'

    # Favicon local
    fav_candidates = sorted(images_dir.glob('favicon.*')) if images_dir.exists() else []
    if not fav_candidates:
        out.append(v('high', 'favicon-asset-missing-local', 'local-assets',
                     {'type': 'file', 'dir': str(images_dir)},
                     f"no favicon.* file in {images_dir}"))
    # Favicon live
    if not snap['site'].get('site_icon'):
        out.append(v('high', 'favicon-not-set-live', 'local-assets',
                     {'type': 'option', 'key': 'site_icon'},
                     "live site_icon option not set — favicon not uploaded",
                     fixable=False))

    # Logo local
    logo_candidates = sorted(images_dir.glob('logo.*')) if images_dir.exists() else []
    if not logo_candidates:
        out.append(v('medium', 'logo-asset-missing-local', 'local-assets',
                     {'type': 'file', 'dir': str(images_dir)},
                     f"no logo.* file in {images_dir}"))

    return out


MD_FILENAME_RE = re.compile(r'^(?:article-\d+-|article\d+_)?(.+)\.md$')


def _md_slug(filename):
    """article10_foo-bar.md → foo-bar, foo.md → foo."""
    m = MD_FILENAME_RE.match(filename)
    return m.group(1) if m else filename.replace('.md', '')


def check_md_html_parity(input_dir, output_dir, snap):
    """For each input md, ensure html output exists and live post exists."""
    out = []
    if not output_dir.exists():
        out.append(v('medium', 'output-dir-missing', 'local-md-html',
                     {'type': 'dir', 'path': str(output_dir)},
                     f"output directory {output_dir} not found — md-html parity skipped"))
        return out

    silos = ['tickets-tours', 'plan-your-visit', 'what-to-see']
    live_slugs = {p['slug'] for p in snap['posts']}

    for silo in silos:
        md_dir   = input_dir  / silo
        html_dir = output_dir / 'l2-articles' / silo
        if not md_dir.exists():
            continue

        md_files = sorted(md_dir.glob('*.md'))
        for md in md_files:
            slug = _md_slug(md.name)
            html_path = html_dir / f"{slug}.html"
            if not html_path.exists():
                out.append(v('high', 'md-without-html', 'local-md-html',
                             {'type': 'slug', 'slug': slug, 'silo': silo},
                             f"{silo}/{md.name} has no matching html at {html_path}",
                             fixable=False))
            if slug not in live_slugs:
                out.append(v('high', 'md-not-published', 'local-md-html',
                             {'type': 'slug', 'slug': slug, 'silo': silo},
                             f"{silo}/{md.name} has no published post on live site",
                             fixable=False))

        # Orphan html: html file with no matching md
        if html_dir.exists():
            html_files  = sorted(html_dir.glob('*.html'))
            md_slugs    = {_md_slug(f.name) for f in md_files}
            for html in html_files:
                slug = html.stem
                if slug not in md_slugs:
                    out.append(v('low', 'html-orphan', 'local-md-html',
                                 {'type': 'slug', 'slug': slug, 'silo': silo},
                                 f"{silo}/{html.name} has no source md file"))

    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('snapshot')
    ap.add_argument('--out', default=None)
    args = ap.parse_args()

    here = Path(__file__).parent
    root = here.parent.parent
    snap = json.load(open(args.snapshot))
    hostname = snap['site']['hostname']

    input_slug = HOSTNAME_TO_INPUT_SLUG.get(hostname)
    if not input_slug:
        print(f"Unknown hostname {hostname} — no input_slug mapping", file=sys.stderr)
        sys.exit(1)

    input_dir  = root / 'input'  / input_slug
    output_dir = root / 'output' / input_slug

    violations = []
    if input_dir.exists():
        violations.extend(check_xlsx_manifest(input_dir, snap))
        violations.extend(check_tickets_md(input_dir, snap))
        violations.extend(check_input_assets(input_dir, snap))
        violations.extend(check_md_html_parity(input_dir, output_dir, snap))
    else:
        violations.append(v('medium', 'input-dir-missing', 'local',
                           {'type': 'dir', 'path': str(input_dir)},
                           f"expected input dir {input_dir} not found"))

    # Same sort order as check.py
    order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
    violations.sort(key=lambda x: (order.get(x['severity'], 9), x['category'], x['id']))

    payload = {
        'snapshot':    str(args.snapshot),
        'hostname':    hostname,
        'input_slug':  input_slug,
        'total':       len(violations),
        'by_severity': {sev: sum(1 for x in violations if x['severity']==sev)
                        for sev in ['critical', 'high', 'medium', 'low']},
        'violations':  violations,
    }

    out_path = args.out
    if not out_path:
        viol_dir = root / 'output' / 'audit' / 'violations'
        viol_dir.mkdir(parents=True, exist_ok=True)
        snap_name = Path(args.snapshot).stem
        out_path = viol_dir / f"{snap_name}.local.json"

    with open(out_path, 'w') as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    print(f"\n=== CHECK_LOCAL: {hostname} (input: {input_slug}) ===", file=sys.stderr)
    print(f"Local violations: {payload['total']}", file=sys.stderr)
    for sev, n in payload['by_severity'].items():
        marker = '!' if n > 0 else ' '
        print(f"  [{marker}] {sev:<9} {n}", file=sys.stderr)
    print(f"\nOutput: {out_path}", file=sys.stderr)


if __name__ == '__main__':
    main()
